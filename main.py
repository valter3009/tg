"""
Telegram –±–æ—Ç GidMeteo –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏ –æ –ø–æ–≥–æ–¥–µ
–∏ –ø–µ—Ä—Å–æ–Ω–∞–ª–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö —Å–æ–≤–µ—Ç–æ–≤ –ø–æ –æ–¥–µ–∂–¥–µ
"""

import telebot
import requests
from telebot import types
import json
import os
import time
import threading
from datetime import datetime, time as dt_time
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, Alignment
from clothes_advice import get_clothing_advice, get_local_time_str, get_time_of_day
from dotenv import load_dotenv
from translations import (
    t, get_user_language, set_user_language,
    SUPPORTED_LANGUAGES, DEFAULT_LANGUAGE
)
from weather_translations import get_wind_description as get_wind_description_i18n

# –ó–∞–≥—Ä—É–∑–∫–∞ –ø–µ—Ä–µ–º–µ–Ω–Ω—ã—Ö –æ–∫—Ä—É–∂–µ–Ω–∏—è
load_dotenv()

# –ù–∞—Å—Ç—Ä–æ–π–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è
logging.basicConfig(
    filename='bot_errors.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# –ù–∞—Å—Ç—Ä–æ–π–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è –¥–ª—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏
activity_logger = logging.getLogger('activity_logger')
activity_handler = logging.FileHandler('activity_tracker.log')
activity_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
activity_logger.addHandler(activity_handler)
activity_logger.setLevel(logging.INFO)

# –ü–æ–ª—É—á–µ–Ω–∏–µ —Ç–æ–∫–µ–Ω–æ–≤ –∏–∑ –ø–µ—Ä–µ–º–µ–Ω–Ω—ã—Ö –æ–∫—Ä—É–∂–µ–Ω–∏—è
TELEGRAM_TOKEN = os.getenv('TELEGRAM_TOKEN')
OPENWEATHER_API_KEY = os.getenv('OPENWEATHER_API_KEY')

if not TELEGRAM_TOKEN:
    raise ValueError("TELEGRAM_TOKEN –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ –ø–µ—Ä–µ–º–µ–Ω–Ω—ã—Ö –æ–∫—Ä—É–∂–µ–Ω–∏—è")
if not OPENWEATHER_API_KEY:
    raise ValueError("OPENWEATHER_API_KEY –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ –ø–µ—Ä–µ–º–µ–Ω–Ω—ã—Ö –æ–∫—Ä—É–∂–µ–Ω–∏—è")

bot = telebot.TeleBot(TELEGRAM_TOKEN)

USER_DATA_FILE = 'user_cities.json'
WEATHER_CACHE_FILE = 'weather_cache.json'
ALL_USERS_FILE = 'all_users.json'  # –ù–æ–≤—ã–π —Ñ–∞–π–ª –¥–ª—è —Ö—Ä–∞–Ω–µ–Ω–∏—è –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
USER_LANGUAGES_FILE = 'user_languages.json'  # –§–∞–π–ª –¥–ª—è —Ö—Ä–∞–Ω–µ–Ω–∏—è —è–∑—ã–∫–æ–≤ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
CACHE_UPDATE_INTERVAL = 3600
ACTIVITY_LOG_FILE = 'bot_activity_log.xlsx'
AUTO_UPDATE_LOG_FILE = 'auto_updates.log'

# –°–ø–∏—Å–æ–∫ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –¥–ª—è —Ä–∞—Å—Å—ã–ª–∫–∏
ADDITIONAL_USERS = [
    471789857, 425748474, 6118023060, 6488366997, 5934413419, 6615704791, 
    1134118381, 1823348752, 6579300547, 5174302370, 1344487460, 7791445179, 
    1276348447, 278283980, 6556640321, 1521820146, 7523695427, 7880850349, 
    832185475, 149653247, 1775572520, 7643533302, 352808232, 7456672724, 
    5969931672, 993675994, 543397394, 5935464436, 1812257315, 6260364812, 
    434939312
]

# –•—Ä–∞–Ω–∏–ª–∏—â–µ –ø–æ—Å–ª–µ–¥–Ω–∏—Ö —Å–æ–æ–±—â–µ–Ω–∏–π —Å timestamp
last_messages = {}

# –°—Ç—Ä—É–∫—Ç—É—Ä–∞ –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —Ö—Ä–∞–Ω–µ–Ω–∏—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏
activity_data = {
    'refresh': {},  # user_id: count
    'cities': {},   # user_id: {city: count}
    'start': {},     # user_id: count - –¥–æ–±–∞–≤–ª–µ–Ω —Å—á–µ—Ç—á–∏–∫ –¥–ª—è /start
    'auto_updates': {'total_sent': 0, 'last_update': None}  # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏–π
}

# –ú—å—é—Ç–µ–∫—Å –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å Excel-—Ñ–∞–π–ª–æ–º –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏
excel_mutex = threading.Lock()

# –§—É–Ω–∫—Ü–∏–∏ –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å—Å–∫–∏–º–∏ –¥–∞–Ω–Ω—ã–º–∏
def load_user_data():
    try:
        if os.path.exists(USER_DATA_FILE):
            with open(USER_DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏ user_data: {e}")
        return {}

def save_user_data(data):
    try:
        with open(USER_DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è user_data: {e}")

# –§—É–Ω–∫—Ü–∏–∏ –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å —è–∑—ã–∫–∞–º–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
def load_user_languages():
    try:
        if os.path.exists(USER_LANGUAGES_FILE):
            with open(USER_LANGUAGES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏ user_languages: {e}")
        return {}

def save_user_languages(data):
    try:
        with open(USER_LANGUAGES_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è user_languages: {e}")

# –ù–æ–≤—ã–µ —Ñ—É–Ω–∫—Ü–∏–∏ –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å–æ –≤—Å–µ–º–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º–∏
def load_all_users():
    try:
        if os.path.exists(ALL_USERS_FILE):
            with open(ALL_USERS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏ all_users: {e}")
        return {}

def save_all_users(data):
    try:
        with open(ALL_USERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è all_users: {e}")

def add_additional_users_to_all_users():
    """–î–æ–±–∞–≤–ª—è–µ—Ç –¥–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –≤ —Å–ø–∏—Å–æ–∫ –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    all_users = load_all_users()
    added_count = 0
    
    for user_id in ADDITIONAL_USERS:
        user_id_str = str(user_id)
        
        if user_id_str not in all_users:
            all_users[user_id_str] = {
                'first_start': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'active': True,
                'source': 'additional_list'  # –ü–æ–º–µ—á–∞–µ–º –∏—Å—Ç–æ—á–Ω–∏–∫
            }
            added_count += 1
        else:
            # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –¥–ª—è —Å—É—â–µ—Å—Ç–≤—É—é—â–∏—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
            all_users[user_id_str]['active'] = True
    
    if added_count > 0:
        save_all_users(all_users)
        logger.info(f"–î–æ–±–∞–≤–ª–µ–Ω–æ {added_count} –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–∑ –¥–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ–≥–æ —Å–ø–∏—Å–∫–∞")
    
    return added_count

def add_user_to_all_users(user_id):
    """–î–æ–±–∞–≤–ª—è–µ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ —Å–ø–∏—Å–æ–∫ –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    all_users = load_all_users()
    user_id_str = str(user_id)
    
    if user_id_str not in all_users:
        all_users[user_id_str] = {
            'first_start': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'active': True,
            'source': 'start_command'  # –ü–æ–º–µ—á–∞–µ–º –∏—Å—Ç–æ—á–Ω–∏–∫
        }
        save_all_users(all_users)
        return True
    else:
        # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏
        all_users[user_id_str]['active'] = True
        save_all_users(all_users)
    return False

def deactivate_user(user_id):
    """–î–µ–∞–∫—Ç–∏–≤–∏—Ä—É–µ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (–Ω–∞–ø—Ä–∏–º–µ—Ä, –µ—Å–ª–∏ –æ–Ω –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–ª –±–æ—Ç–∞)"""
    all_users = load_all_users()
    user_id_str = str(user_id)
    
    if user_id_str in all_users:
        all_users[user_id_str]['active'] = False
        save_all_users(all_users)

def load_weather_cache():
    try:
        if os.path.exists(WEATHER_CACHE_FILE):
            with open(WEATHER_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏ weather_cache: {e}")
        return {}

def save_weather_cache(data):
    try:
        with open(WEATHER_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è weather_cache: {e}")

def get_cached_weather(city, weather_cache):
    return weather_cache.get(city.lower())

def update_cached_weather(city, weather_data, weather_cache):
    weather_cache[city.lower()] = weather_data
    save_weather_cache(weather_cache)

def get_user_cities(user_id, user_data):
    user_id_str = str(user_id)
    return user_data.get(user_id_str, [])

def add_user_city(user_id, city, user_data):
    user_id_str = str(user_id)
    
    if user_id_str not in user_data:
        user_data[user_id_str] = []
    
    if city not in user_data[user_id_str]:
        user_data[user_id_str].append(city)
        save_user_data(user_data)
        return True
    return False

def remove_user_city(user_id, city, user_data):
    user_id_str = str(user_id)
    
    if user_id_str in user_data and city in user_data[user_id_str]:
        user_data[user_id_str].remove(city)
        save_user_data(user_data)
        return True
    return False

def get_time_of_day_emoji(timezone_offset):
    """
    –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç —ç–º–æ–¥–∑–∏ –≤—Ä–µ–º–µ–Ω–∏ —Å—É—Ç–æ–∫ –Ω–∞ –æ—Å–Ω–æ–≤–µ –ª–æ–∫–∞–ª—å–Ω–æ–≥–æ –≤—Ä–µ–º–µ–Ω–∏

    Args:
        timezone_offset (int): –°–º–µ—â–µ–Ω–∏–µ —á–∞—Å–æ–≤–æ–≥–æ –ø–æ—è—Å–∞ –≤ —Å–µ–∫—É–Ω–¥–∞—Ö –æ—Ç UTC

    Returns:
        str: –≠–º–æ–¥–∑–∏ –≤—Ä–µ–º–µ–Ω–∏ —Å—É—Ç–æ–∫
    """
    time_of_day = get_time_of_day(timezone_offset)

    if time_of_day == '—É—Ç—Ä–æ':
        return 'üåÖ'  # —Ä–∞—Å—Å–≤–µ—Ç
    elif time_of_day == '–¥–µ–Ω—å':
        return '‚òÄÔ∏è'  # —Å–æ–ª–Ω—Ü–µ
    elif time_of_day == '–≤–µ—á–µ—Ä':
        return 'üåÜ'  # –∑–∞–∫–∞—Ç
    else:  # –Ω–æ—á—å
        return 'üåô'  # –ª—É–Ω–∞

def get_weather_emoji(description):
    description = description.lower()

    if any(word in description for word in ['—è—Å–Ω–æ', '—á–∏—Å—Ç–æ–µ –Ω–µ–±–æ', '–±–µ–∑–æ–±–ª–∞—á–Ω–æ']):
        return '‚òÄÔ∏è'
    elif any(word in description for word in ['–æ–±–ª–∞—á–Ω–æ —Å –ø—Ä–æ—è—Å–Ω–µ–Ω–∏—è–º–∏', '–ø–µ—Ä–µ–º–µ–Ω–Ω–∞—è –æ–±–ª–∞—á–Ω–æ—Å—Ç—å']):
        return '‚õÖ'
    elif any(word in description for word in ['–æ–±–ª–∞—á–Ω–æ', '–ø–∞—Å–º—É—Ä–Ω–æ']):
        return '‚òÅÔ∏è'
    elif any(word in description for word in ['–¥–æ–∂–¥—å', '–ª–∏–≤–µ–Ω—å']):
        return 'üåßÔ∏è'
    elif any(word in description for word in ['–≥—Ä–æ–∑–∞']):
        return '‚õàÔ∏è'
    elif any(word in description for word in ['—Å–Ω–µ–≥', '—Å–Ω–µ–≥–æ–ø–∞–¥']):
        return '‚ùÑÔ∏è'
    elif any(word in description for word in ['—Ç—É–º–∞–Ω', '–º–≥–ª–∞']):
        return 'üå´Ô∏è'
    else:
        return 'üå¶Ô∏è'

def get_weather_data(city, weather_cache):
    url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&units=metric&lang=ru&appid={OPENWEATHER_API_KEY}'

    try:
        response = requests.get(url)
        response.raise_for_status()
        weather_data = response.json()

        if weather_data.get('cod') == '404':
            return None

        temperature = round(weather_data['main']['temp'], 1)
        weather_description = weather_data['weather'][0]['description']
        wind_speed = int(round(weather_data['wind']['speed']))
        timezone_offset = weather_data.get('timezone', 0)  # –°–º–µ—â–µ–Ω–∏–µ —á–∞—Å–æ–≤–æ–≥–æ –ø–æ—è—Å–∞ –≤ —Å–µ–∫—É–Ω–¥–∞—Ö

        weather_emoji = get_weather_emoji(weather_description)

        result = {
            'temp': temperature,
            'emoji': weather_emoji,
            'description': weather_description,
            'wind_speed': wind_speed,
            'timezone': timezone_offset,
            'updated_at': int(time.time())
        }

        update_cached_weather(city, result, weather_cache)
        return result

    except Exception as e:
        logger.error(f'–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ –ø–æ–≥–æ–¥—ã –¥–ª—è {city}: {e}')
        return None

def format_date_time(timestamp):
    return datetime.fromtimestamp(timestamp).strftime('%d.%m.%Y %H:%M:%S')

def format_local_date_time(timezone_offset):
    """
    –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç –ª–æ–∫–∞–ª—å–Ω—É—é –¥–∞—Ç—É –∏ –≤—Ä–µ–º—è –≥–æ—Ä–æ–¥–∞

    Args:
        timezone_offset (int): –°–º–µ—â–µ–Ω–∏–µ —á–∞—Å–æ–≤–æ–≥–æ –ø–æ—è—Å–∞ –≤ —Å–µ–∫—É–Ω–¥–∞—Ö –æ—Ç UTC

    Returns:
        str: –î–∞—Ç–∞ –∏ –≤—Ä–µ–º—è –≤ —Ñ–æ—Ä–º–∞—Ç–µ DD.MM.YYYY HH:MM:SS
    """
    from datetime import datetime, timezone, timedelta

    utc_time = datetime.now(timezone.utc)
    local_time = utc_time + timedelta(seconds=timezone_offset)
    return local_time.strftime('%d.%m.%Y %H:%M:%S')

def update_weather_cache(weather_cache, user_data):
    all_cities = set()
    
    for user_id, cities in user_data.items():
        for city in cities:
            all_cities.add(city)
    
    for city in all_cities:
        try:
            get_weather_data(city, weather_cache)
            time.sleep(1)
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–∏ –ø–æ–≥–æ–¥—ã –¥–ª—è {city}: {e}")

def create_cities_keyboard(user_id, user_data, weather_cache, user_languages):
    markup = types.InlineKeyboardMarkup(row_width=1)
    user_cities = get_user_cities(user_id, user_data)
    language = get_user_language(user_id, user_languages)

    for city in user_cities:
        cached_weather = get_cached_weather(city, weather_cache)
        current_time = int(time.time())

        if not cached_weather or (current_time - cached_weather.get('updated_at', 0)) > 3600:
            weather_data = get_weather_data(city, weather_cache)
            if weather_data:
                temp_str = f"+{weather_data['temp']}" if weather_data['temp'] > 0 else f"{weather_data['temp']}"
                # –ü–æ–ª—É—á–∞–µ–º —ç–º–æ–¥–∑–∏ –≤—Ä–µ–º–µ–Ω–∏ —Å—É—Ç–æ–∫
                time_emoji = get_time_of_day_emoji(weather_data.get('timezone', 0))
                wind_label = "üí®" if language == 'ru' else "üí®"
                button_text = f"{weather_data['emoji']} {city} {time_emoji} {temp_str}¬∞C {wind_label}{weather_data['wind_speed']}–º/—Å"
            else:
                button_text = city
        else:
            temp_str = f"+{cached_weather['temp']}" if cached_weather['temp'] > 0 else f"{cached_weather['temp']}"
            # –ü–æ–ª—É—á–∞–µ–º —ç–º–æ–¥–∑–∏ –≤—Ä–µ–º–µ–Ω–∏ —Å—É—Ç–æ–∫ –∏–∑ –∫–µ—à–∞
            time_emoji = get_time_of_day_emoji(cached_weather.get('timezone', 0))
            wind_label = "üí®" if language == 'ru' else "üí®"
            button_text = f"{cached_weather['emoji']} {city} {time_emoji} {temp_str}¬∞C {wind_label}{cached_weather['wind_speed']}–º/—Å"

        markup.add(types.InlineKeyboardButton(text=button_text, callback_data=f"city_{city}"))

    # –î–æ–±–∞–≤–ª—è–µ–º –∫–Ω–æ–ø–∫—É "–û–±–Ω–æ–≤–∏—Ç—å", —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –µ—Å—Ç—å –≥–æ—Ä–æ–¥–∞
    if user_cities:
        refresh_text = t('refresh_button', language)
        markup.add(types.InlineKeyboardButton(text=refresh_text, callback_data="refresh"))

    return markup

def send_new_message(chat_id, text, markup=None, parse_mode='Markdown'):
    try:
        sent_msg = bot.send_message(
            chat_id,
            text,
            reply_markup=markup,
            parse_mode=parse_mode
        )
        last_messages[chat_id] = {
            'message_id': sent_msg.message_id,
            'timestamp': time.time()
        }
        return sent_msg
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–ø—Ä–∞–≤–∫–µ –Ω–æ–≤–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è: {e}")
        return None

def update_message(chat_id, message_id, text, markup=None, parse_mode='Markdown'):
    try:
        if message_id:
            try:
                return bot.edit_message_text(
                    text=text,
                    chat_id=chat_id,
                    message_id=message_id,
                    reply_markup=markup,
                    parse_mode=parse_mode
                )
            except telebot.apihelper.ApiTelegramException as e:
                if "message to edit not found" in str(e).lower():
                    logger.info(f"–°–æ–æ–±—â–µ–Ω–∏–µ {message_id} –≤ —á–∞—Ç–µ {chat_id} –Ω–µ –Ω–∞–π–¥–µ–Ω–æ, –æ—Ç–ø—Ä–∞–≤–ª—è–µ–º –Ω–æ–≤–æ–µ")
                    if chat_id in last_messages:
                        del last_messages[chat_id]
                    return send_new_message(chat_id, text, markup, parse_mode)
                raise
        return send_new_message(chat_id, text, markup, parse_mode)
    
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–∏ —Å–æ–æ–±—â–µ–Ω–∏—è: {e}")
        return send_new_message(chat_id, text, markup, parse_mode)

def send_welcome_message(chat_id, user_data, weather_cache, user_languages, message_id=None, force_new_message=False):
    language = get_user_language(chat_id, user_languages)
    markup = create_cities_keyboard(chat_id, user_data, weather_cache, user_languages)

    user_cities = get_user_cities(chat_id, user_data)
    cities_weather_text = []

    for city in user_cities:
        cached_weather = get_cached_weather(city, weather_cache)
        if cached_weather:
            temp_str = f"+{cached_weather['temp']}" if cached_weather['temp'] > 0 else f"{cached_weather['temp']}"
            city_text = f"{cached_weather['emoji']} {city} {temp_str}¬∞C"
            cities_weather_text.append(city_text)
        else:
            cities_weather_text.append(city)

    if user_cities:
        cities_str = "\n".join(cities_weather_text)
        welcome_text = t('welcome_text_with_cities', language, cities=cities_str)
    else:
        welcome_text = t('welcome_text', language)
    
    if force_new_message:
        # –°–Ω–∞—á–∞–ª–∞ —É–¥–∞–ª—è–µ–º –ø—Ä–µ–¥—ã–¥—É—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
        try:
            if message_id:
                bot.delete_message(chat_id, message_id)
        except Exception as e:
            logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å –ø—Ä–µ–¥—ã–¥—É—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ: {e}")
        
        # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –Ω–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
        sent_message = send_new_message(chat_id, welcome_text, markup)
    else:
        # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
        msg_info = last_messages.get(chat_id, {})
        msg_id = message_id if message_id else msg_info.get('message_id')
        
        sent_message = update_message(chat_id, msg_id, welcome_text, markup)
    
    if sent_message and hasattr(sent_message, 'message_id'):
        last_messages[chat_id] = {
            'message_id': sent_message.message_id,
            'timestamp': time.time()
        }

def send_reminder_message(chat_id, user_languages):
    """–û—Ç–ø—Ä–∞–≤–ª—è–µ—Ç –Ω–∞–ø–æ–º–∏–Ω–∞–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º –±–µ–∑ –≥–æ—Ä–æ–¥–æ–≤"""
    language = get_user_language(chat_id, user_languages)
    reminder_text = t('reminder_text', language)
    
    try:
        sent_msg = bot.send_message(chat_id, reminder_text)
        last_messages[chat_id] = {
            'message_id': sent_msg.message_id,
            'timestamp': time.time()
        }
        return True
    except telebot.apihelper.ApiTelegramException as e:
        if "bot was blocked by the user" in str(e).lower():
            deactivate_user(chat_id)
            logger.info(f"–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å {chat_id} –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–ª –±–æ—Ç–∞. –î–µ–∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω.")
            return False
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–ø—Ä–∞–≤–∫–µ –Ω–∞–ø–æ–º–∏–Ω–∞–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é {chat_id}: {e}")
        return False
    
# –§—É–Ω–∫—Ü–∏–∏ –¥–ª—è –æ—Ç—Å–ª–µ–∂–∏–≤–∞–Ω–∏—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏
def initialize_excel_file():
    """–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ—Ç Excel-—Ñ–∞–π–ª –¥–ª—è –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è, –µ—Å–ª–∏ –æ–Ω –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç"""
    try:
        if not os.path.exists(ACTIVITY_LOG_FILE):
            # –°–æ–∑–¥–∞–µ–º DataFrame –¥–ª—è –≤–∫–ª–∞–¥–∫–∏ —Å –∫–Ω–æ–ø–∫–æ–π "–û–±–Ω–æ–≤–∏—Ç—å"
            df_refresh = pd.DataFrame(columns=['User ID', 'Count', 'Last Update'])
            
            # –°–æ–∑–¥–∞–µ–º DataFrame –¥–ª—è –≤–∫–ª–∞–¥–∫–∏ —Å –≥–æ—Ä–æ–¥–∞–º–∏
            df_cities = pd.DataFrame(columns=['User ID', 'City', 'Count', 'Last Update'])
            
            # –°–æ–∑–¥–∞–µ–º DataFrame –¥–ª—è –≤–∫–ª–∞–¥–∫–∏ —Å –∫–æ–º–∞–Ω–¥–æ–π /start
            df_start = pd.DataFrame(columns=['User ID', 'Count', 'Last Update'])
            
            # –°–æ–∑–¥–∞–µ–º DataFrame –¥–ª—è –≤–∫–ª–∞–¥–∫–∏ —Å –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏—è–º–∏
            df_auto_updates = pd.DataFrame(columns=['Date', 'Time', 'Total Attempts', 'Messages Sent', 'Users With Cities', 'Users Without Cities', 'Blocked Users', 'Errors'])
            
            # –°–æ–∑–¥–∞–µ–º writer –¥–ª—è –∑–∞–ø–∏—Å–∏ –≤ Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl') as writer:
                df_refresh.to_excel(writer, sheet_name='Refresh Button', index=False)
                df_cities.to_excel(writer, sheet_name='City Buttons', index=False)
                df_start.to_excel(writer, sheet_name='Start Command', index=False)
                df_auto_updates.to_excel(writer, sheet_name='Auto Updates', index=False)
            
            activity_logger.info(f"–°–æ–∑–¥–∞–Ω –Ω–æ–≤—ã–π —Ñ–∞–π–ª –ª–æ–≥–æ–≤ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏: {ACTIVITY_LOG_FILE}")
        else:
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ñ–∞–π–ª –Ω–∞ –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ—Å—Ç—å
            try:
                workbook = load_workbook(ACTIVITY_LOG_FILE)
                
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –ª–∏—Å—Ç–∞ –¥–ª—è –∫–æ–º–∞–Ω–¥—ã /start
                if 'Start Command' not in workbook.sheetnames:
                    df_start = pd.DataFrame(columns=['User ID', 'Count', 'Last Update'])
                    with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a') as writer:
                        df_start.to_excel(writer, sheet_name='Start Command', index=False)
                    activity_logger.info(f"–î–æ–±–∞–≤–ª–µ–Ω –ª–∏—Å—Ç 'Start Command' –≤ —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π —Ñ–∞–π–ª –ª–æ–≥–æ–≤")
                
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –ª–∏—Å—Ç–∞ –¥–ª—è –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏–π
                if 'Auto Updates' not in workbook.sheetnames:
                    df_auto_updates = pd.DataFrame(columns=['Date', 'Time', 'Total Attempts', 'Messages Sent', 'Users With Cities', 'Users Without Cities', 'Blocked Users', 'Errors'])
                    with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a') as writer:
                        df_auto_updates.to_excel(writer, sheet_name='Auto Updates', index=False)
                    activity_logger.info(f"–î–æ–±–∞–≤–ª–µ–Ω –ª–∏—Å—Ç 'Auto Updates' –≤ —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π —Ñ–∞–π–ª –ª–æ–≥–æ–≤")
                
                activity_logger.info(f"–°—É—â–µ—Å—Ç–≤—É—é—â–∏–π —Ñ–∞–π–ª –ª–æ–≥–æ–≤ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –∫–æ—Ä—Ä–µ–∫—Ç–µ–Ω: {ACTIVITY_LOG_FILE}")
            except InvalidFileException:
                activity_logger.error(f"–§–∞–π–ª –ª–æ–≥–æ–≤ –ø–æ–≤—Ä–µ–∂–¥–µ–Ω, —Å–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π: {ACTIVITY_LOG_FILE}")
                os.remove(ACTIVITY_LOG_FILE)
                initialize_excel_file()
    except Exception as e:
        activity_logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏–∏ —Ñ–∞–π–ª–∞ –ª–æ–≥–æ–≤: {e}")

def update_refresh_log(user_id):
    """–û–±–Ω–æ–≤–ª—è–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –Ω–∞–∂–∞—Ç–∏–∏ –∫–Ω–æ–ø–∫–∏ '–û–±–Ω–æ–≤–∏—Ç—å'"""
    try:
        with excel_mutex:
            # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –≤ –ø–∞–º—è—Ç–∏
            if user_id in activity_data['refresh']:
                activity_data['refresh'][user_id] += 1
            else:
                activity_data['refresh'][user_id] = 1
            
            # –¢–µ–∫—É—â–∞—è –¥–∞—Ç–∞ –∏ –≤—Ä–µ–º—è
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ Excel
            df = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Refresh Button')
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —É–∂–µ –∑–∞–ø–∏—Å—å –¥–ª—è –¥–∞–Ω–Ω–æ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
            user_row = df[df['User ID'] == user_id]
            
            if len(user_row) > 0:
                # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â—É—é –∑–∞–ø–∏—Å—å
                df.loc[df['User ID'] == user_id, 'Count'] = activity_data['refresh'][user_id]
                df.loc[df['User ID'] == user_id, 'Last Update'] = current_time
            else:
                # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –∑–∞–ø–∏—Å—å
                new_row = pd.DataFrame({
                    'User ID': [user_id],
                    'Count': [activity_data['refresh'][user_id]],
                    'Last Update': [current_time]
                })
                df = pd.concat([df, new_row], ignore_index=True)
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –¥–∞–Ω–Ω—ã–µ –æ–±—Ä–∞—Ç–Ω–æ –≤ Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Refresh Button', index=False)
                
            activity_logger.info(f"–û–±–Ω–æ–≤–ª–µ–Ω –ª–æ–≥ –Ω–∞–∂–∞—Ç–∏—è –∫–Ω–æ–ø–∫–∏ '–û–±–Ω–æ–≤–∏—Ç—å' –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user_id}")
    except Exception as e:
        activity_logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–∏ –ª–æ–≥–∞ –Ω–∞–∂–∞—Ç–∏—è –∫–Ω–æ–ø–∫–∏ '–û–±–Ω–æ–≤–∏—Ç—å': {e}")

def update_auto_updates_log(sent_count, with_cities_count, without_cities_count, blocked_count=0, error_count=0, total_attempts=0):
    """–û–±–Ω–æ–≤–ª—è–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ–± –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏—Ö –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è—Ö"""
    try:
        with excel_mutex:
            # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –≤ –ø–∞–º—è—Ç–∏
            activity_data['auto_updates']['total_sent'] += sent_count
            activity_data['auto_updates']['last_update'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # –¢–µ–∫—É—â–∞—è –¥–∞—Ç–∞ –∏ –≤—Ä–µ–º—è
            current_date = datetime.now().strftime('%Y-%m-%d')
            current_time = datetime.now().strftime('%H:%M:%S')
            
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ Excel
            df = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Auto Updates')
            
            # –î–æ–±–∞–≤–ª—è–µ–º –Ω–æ–≤—É—é –∑–∞–ø–∏—Å—å
            new_row = pd.DataFrame({
                'Date': [current_date],
                'Time': [current_time],
                'Total Attempts': [total_attempts],
                'Messages Sent': [sent_count],
                'Users With Cities': [with_cities_count],
                'Users Without Cities': [without_cities_count],
                'Blocked Users': [blocked_count],
                'Errors': [error_count]
            })
            df = pd.concat([df, new_row], ignore_index=True)
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –¥–∞–Ω–Ω—ã–µ –æ–±—Ä–∞—Ç–Ω–æ –≤ Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Auto Updates', index=False)
                
            activity_logger.info(f"–û–±–Ω–æ–≤–ª–µ–Ω –ª–æ–≥ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏—Ö –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π: –ø–æ–ø—ã—Ç–æ–∫ {total_attempts}, –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω–æ {sent_count} —Å–æ–æ–±—â–µ–Ω–∏–π (—Å –≥–æ—Ä–æ–¥–∞–º–∏: {with_cities_count}, –±–µ–∑ –≥–æ—Ä–æ–¥–æ–≤: {without_cities_count}, –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω–æ: {blocked_count}, –æ—à–∏–±–æ–∫: {error_count})")
    except Exception as e:
        activity_logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–∏ –ª–æ–≥–∞ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏—Ö –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π: {e}")

def auto_update_users():
    """–ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –æ–±–Ω–æ–≤–ª—è–µ—Ç —Å–æ–æ–±—â–µ–Ω–∏—è –¥–ª—è –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π 4 —Ä–∞–∑–∞ –≤ —Å—É—Ç–∫–∏ –ø–æ –ª–æ–∫–∞–ª—å–Ω–æ–º—É –≤—Ä–µ–º–µ–Ω–∏ –ø–µ—Ä–≤–æ–≥–æ –≥–æ—Ä–æ–¥–∞"""
    # –°–ª–æ–≤–∞—Ä—å –¥–ª—è –æ—Ç—Å–ª–µ–∂–∏–≤–∞–Ω–∏—è –ø–æ—Å–ª–µ–¥–Ω–∏—Ö –æ—Ç–ø—Ä–∞–≤–æ–∫ –∫–∞–∂–¥–æ–º—É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
    last_sent = {}  # {user_id: last_update_hour}

    while True:
        try:
            from datetime import datetime, timezone, timedelta

            # –¢–µ–∫—É—â–µ–µ –≤—Ä–µ–º—è UTC
            now_utc = datetime.now(timezone.utc)
            current_minute = now_utc.minute

            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ç–æ–ª—å–∫–æ –≤ –ø–µ—Ä–≤—ã–µ 3 –º–∏–Ω—É—Ç—ã –∫–∞–∂–¥–æ–≥–æ —á–∞—Å–∞
            if current_minute <= 2:
                user_data = load_user_data()
                weather_cache = load_weather_cache()
                all_users = load_all_users()

                # –°—á–µ—Ç—á–∏–∫–∏ –¥–ª—è –ø–æ–¥—Ä–æ–±–Ω–æ–π —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏
                total_attempts = 0
                sent_with_cities = 0
                sent_without_cities = 0
                blocked_count = 0
                error_count = 0
                total_sent = 0

                # –í—Ä–µ–º–µ–Ω–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π –ø–æ –ª–æ–∫–∞–ª—å–Ω–æ–º—É –≤—Ä–µ–º–µ–Ω–∏
                update_hours = [7, 12, 18, 22]

                # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –≤—Å–µ—Ö –∞–∫—Ç–∏–≤–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
                for user_id_str, user_info in all_users.items():
                    if not user_info.get('active', True):
                        continue

                    try:
                        user_id = int(user_id_str)
                        user_cities = get_user_cities(user_id, user_data)

                        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º timezone –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
                        if user_cities:
                            # –ü–æ–ª—É—á–∞–µ–º timezone –ø–µ—Ä–≤–æ–≥–æ –≥–æ—Ä–æ–¥–∞
                            first_city = user_cities[0]
                            weather_data = get_cached_weather(first_city, weather_cache)
                            if not weather_data:
                                weather_data = get_weather_data(first_city, weather_cache)

                            user_timezone = weather_data.get('timezone', 10800) if weather_data else 10800  # –ü–æ —É–º–æ–ª—á–∞–Ω–∏—é –ú–æ—Å–∫–≤–∞ UTC+3
                        else:
                            # –ï—Å–ª–∏ –Ω–µ—Ç –≥–æ—Ä–æ–¥–æ–≤, –∏—Å–ø–æ–ª—å–∑—É–µ–º –ú–æ—Å–∫–æ–≤—Å–∫–æ–µ –≤—Ä–µ–º—è
                            user_timezone = 10800  # –ú–æ—Å–∫–≤–∞ UTC+3

                        # –í—ã—á–∏—Å–ª—è–µ–º –ª–æ–∫–∞–ª—å–Ω–æ–µ –≤—Ä–µ–º—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
                        local_time = now_utc + timedelta(seconds=user_timezone)
                        local_hour = local_time.hour

                        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω—É–∂–Ω–æ –ª–∏ –æ—Ç–ø—Ä–∞–≤–ª—è—Ç—å –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ
                        should_update = local_hour in update_hours
                        already_sent = last_sent.get(user_id) == local_hour

                        if should_update and not already_sent:
                            total_attempts += 1

                            if user_cities:
                                # –û–±–Ω–æ–≤–ª—è–µ–º –∫—ç—à –ø–æ–≥–æ–¥—ã –ø–µ—Ä–µ–¥ –æ—Ç–ø—Ä–∞–≤–∫–æ–π
                                update_weather_cache(weather_cache, user_data)

                                # –ó–∞–≥—Ä—É–∂–∞–µ–º —è–∑—ã–∫–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
                                user_languages = load_user_languages()

                                # –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —Å –≥–æ—Ä–æ–¥–∞–º–∏ - –æ—Ç–ø—Ä–∞–≤–ª—è–µ–º –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ –ø–æ–≥–æ–¥—ã
                                msg_info = last_messages.get(user_id, {})
                                message_id = msg_info.get('message_id')

                                send_welcome_message(user_id, user_data, weather_cache, user_languages, message_id, force_new_message=True)
                                sent_with_cities += 1
                                total_sent += 1
                            else:
                                # –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –±–µ–∑ –≥–æ—Ä–æ–¥–æ–≤ - –æ—Ç–ø—Ä–∞–≤–ª—è–µ–º –Ω–∞–ø–æ–º–∏–Ω–∞–Ω–∏–µ
                                user_languages = load_user_languages()
                                if send_reminder_message(user_id, user_languages):
                                    sent_without_cities += 1
                                    total_sent += 1

                            # –ó–∞–ø–æ–º–∏–Ω–∞–µ–º, —á—Ç–æ –æ—Ç–ø—Ä–∞–≤–∏–ª–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ –≤ —ç—Ç–æ—Ç —á–∞—Å
                            last_sent[user_id] = local_hour

                            # –ù–µ–±–æ–ª—å—à–∞—è –∑–∞–¥–µ—Ä–∂–∫–∞ –º–µ–∂–¥—É –æ—Ç–ø—Ä–∞–≤–∫–∞–º–∏
                            time.sleep(0.1)

                    except telebot.apihelper.ApiTelegramException as e:
                        if "bot was blocked by the user" in str(e).lower():
                            # –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–ª –±–æ—Ç–∞
                            deactivate_user(user_id_str)
                            blocked_count += 1
                            logger.info(f"–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å {user_id_str} –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–ª –±–æ—Ç–∞. –î–µ–∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω.")
                        else:
                            error_count += 1
                            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–ø—Ä–∞–≤–∫–µ –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é {user_id_str}: {e}")
                    except Exception as e:
                        error_count += 1
                        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–ø—Ä–∞–≤–∫–µ –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é {user_id_str}: {e}")

                # –õ–æ–≥–∏—Ä—É–µ–º –ø–æ–¥—Ä–æ–±–Ω—É—é —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É
                if total_sent > 0:
                    update_auto_updates_log(total_sent, sent_with_cities, sent_without_cities, blocked_count, error_count, total_attempts)

                    # –ü–æ–¥—Ä–æ–±–Ω—ã–π –ª–æ–≥
                    success_rate = (total_sent / total_attempts) * 100 if total_attempts > 0 else 0
                    logger.info(f"–ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–æ–µ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ –≤ {now_utc.strftime('%H:%M:%S')} UTC:")
                    logger.info(f"  - –í—Å–µ–≥–æ –ø–æ–ø—ã—Ç–æ–∫: {total_attempts}")
                    logger.info(f"  - –£—Å–ø–µ—à–Ω–æ –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω–æ: {total_sent} ({success_rate:.1f}%)")
                    logger.info(f"  - –° –≥–æ—Ä–æ–¥–∞–º–∏: {sent_with_cities}")
                    logger.info(f"  - –ë–µ–∑ –≥–æ—Ä–æ–¥–æ–≤: {sent_without_cities}")
                    logger.info(f"  - –ó–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–ª–∏ –±–æ—Ç–∞: {blocked_count}")
                    logger.info(f"  - –û—à–∏–±–∫–∏ –æ—Ç–ø—Ä–∞–≤–∫–∏: {error_count}")

                # –ñ–¥–µ–º –¥–æ —Å–ª–µ–¥—É—é—â–µ–π –º–∏–Ω—É—Ç—ã
                time.sleep(60)
            else:
                # –ï—Å–ª–∏ –Ω–µ –≤—Ä–µ–º—è –ø—Ä–æ–≤–µ—Ä–∫–∏, –∂–¥–µ–º 30 —Å–µ–∫—É–Ω–¥
                time.sleep(30)

        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –≤ auto_update_users: {e}")
            time.sleep(300)  # 5 –º–∏–Ω—É—Ç –≤ —Å–ª—É—á–∞–µ –æ—à–∏–±–∫–∏

def update_start_log(user_id):
    """–û–±–Ω–æ–≤–ª—è–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–∏ –∫–æ–º–∞–Ω–¥—ã /start"""
    try:
        with excel_mutex:
            # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –≤ –ø–∞–º—è—Ç–∏
            if user_id in activity_data['start']:
                activity_data['start'][user_id] += 1
            else:
                activity_data['start'][user_id] = 1
            
            # –¢–µ–∫—É—â–∞—è –¥–∞—Ç–∞ –∏ –≤—Ä–µ–º—è
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ Excel
            df = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Start Command')
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —É–∂–µ –∑–∞–ø–∏—Å—å –¥–ª—è –¥–∞–Ω–Ω–æ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
            user_row = df[df['User ID'] == user_id]
            
            if len(user_row) > 0:
                # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â—É—é –∑–∞–ø–∏—Å—å
                df.loc[df['User ID'] == user_id, 'Count'] = activity_data['start'][user_id]
                df.loc[df['User ID'] == user_id, 'Last Update'] = current_time
            else:
                # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –∑–∞–ø–∏—Å—å
                new_row = pd.DataFrame({
                    'User ID': [user_id],
                    'Count': [activity_data['start'][user_id]],
                    'Last Update': [current_time]
                })
                df = pd.concat([df, new_row], ignore_index=True)
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –¥–∞–Ω–Ω—ã–µ –æ–±—Ä–∞—Ç–Ω–æ –≤ Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Start Command', index=False)
                
            activity_logger.info(f"–û–±–Ω–æ–≤–ª–µ–Ω –ª–æ–≥ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è –∫–æ–º–∞–Ω–¥—ã /start –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user_id}")
    except Exception as e:
        activity_logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–∏ –ª–æ–≥–∞ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è –∫–æ–º–∞–Ω–¥—ã /start: {e}")

def update_city_log(user_id, city):
    """–û–±–Ω–æ–≤–ª—è–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –Ω–∞–∂–∞—Ç–∏–∏ –∫–Ω–æ–ø–∫–∏ —Å –Ω–∞–∑–≤–∞–Ω–∏–µ–º –≥–æ—Ä–æ–¥–∞"""
    try:
        with excel_mutex:
            # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –≤ –ø–∞–º—è—Ç–∏
            if user_id not in activity_data['cities']:
                activity_data['cities'][user_id] = {}
                
            if city in activity_data['cities'][user_id]:
                activity_data['cities'][user_id][city] += 1
            else:
                activity_data['cities'][user_id][city] = 1
            
            # –¢–µ–∫—É—â–∞—è –¥–∞—Ç–∞ –∏ –≤—Ä–µ–º—è
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ Excel
            df = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='City Buttons')
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —É–∂–µ –∑–∞–ø–∏—Å—å –¥–ª—è –¥–∞–Ω–Ω–æ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –∏ –≥–æ—Ä–æ–¥–∞
            user_city_row = df[(df['User ID'] == user_id) & (df['City'] == city)]
            
            if len(user_city_row) > 0:
                # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â—É—é –∑–∞–ø–∏—Å—å
                df.loc[(df['User ID'] == user_id) & (df['City'] == city), 'Count'] = activity_data['cities'][user_id][city]
                df.loc[(df['User ID'] == user_id) & (df['City'] == city), 'Last Update'] = current_time
            else:
                # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –∑–∞–ø–∏—Å—å
                new_row = pd.DataFrame({
                    'User ID': [user_id],
                    'City': [city],
                    'Count': [activity_data['cities'][user_id][city]],
                    'Last Update': [current_time]
                })
                df = pd.concat([df, new_row], ignore_index=True)
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –¥–∞–Ω–Ω—ã–µ –æ–±—Ä–∞—Ç–Ω–æ –≤ Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='City Buttons', index=False)
                
            activity_logger.info(f"–û–±–Ω–æ–≤–ª–µ–Ω –ª–æ–≥ –Ω–∞–∂–∞—Ç–∏—è –∫–Ω–æ–ø–∫–∏ –≥–æ—Ä–æ–¥–∞ {city} –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user_id}")
    except Exception as e:
        activity_logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–∏ –ª–æ–≥–∞ –Ω–∞–∂–∞—Ç–∏—è –∫–Ω–æ–ø–∫–∏ –≥–æ—Ä–æ–¥–∞: {e}")

def generate_activity_report():
    """–ì–µ–Ω–µ—Ä–∏—Ä—É–µ—Ç –æ—Ç—á–µ—Ç –æ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    try:
        with excel_mutex:
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ Excel
            df_refresh = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Refresh Button')
            df_cities = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='City Buttons')
            df_start = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Start Command')
            df_auto_updates = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Auto Updates')
            
            # –û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
            total_refresh_clicks = df_refresh['Count'].sum() if not df_refresh.empty else 0
            total_city_clicks = df_cities['Count'].sum() if not df_cities.empty else 0
            total_start_cmds = df_start['Count'].sum() if not df_start.empty else 0
            total_auto_updates = df_auto_updates['Users Count'].sum() if not df_auto_updates.empty else 0
            
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –æ –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è—Ö
            all_users = load_all_users()
            active_users = sum(1 for user in all_users.values() if user.get('active', True))
            
            # –¢–æ–ø-5 –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –ø–æ –Ω–∞–∂–∞—Ç–∏—é –∫–Ω–æ–ø–∫–∏ "–û–±–Ω–æ–≤–∏—Ç—å"
            top_refresh_users = df_refresh.sort_values('Count', ascending=False).head(5) if not df_refresh.empty else pd.DataFrame()
            
            # –¢–æ–ø-5 –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –ø–æ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—é –∫–æ–º–∞–Ω–¥—ã /start
            top_start_users = df_start.sort_values('Count', ascending=False).head(5) if not df_start.empty else pd.DataFrame()
            
            # –¢–æ–ø-5 —Å–∞–º—ã—Ö –ø–æ–ø—É–ª—è—Ä–Ω—ã—Ö –≥–æ—Ä–æ–¥–æ–≤
            if not df_cities.empty and 'City' in df_cities.columns:
                top_cities = df_cities.groupby('City')['Count'].sum().reset_index().sort_values('Count', ascending=False).head(5)
            else:
                top_cities = pd.DataFrame()
            
            # –ü–æ—Å–ª–µ–¥–Ω–∏–µ 5 –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏—Ö –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π
            if not df_auto_updates.empty:
                last_updates = df_auto_updates.tail(5)
                last_updates_str = "\n".join([
                    f"{row['Date']} {row['Time']} - –ü–æ–ø—ã—Ç–æ–∫: {row.get('Total Attempts', 'N/A')}, "
                    f"–û—Ç–ø—Ä–∞–≤–ª–µ–Ω–æ: {row.get('Messages Sent', row.get('Users Count', 'N/A'))}, "
                    f"–° –≥–æ—Ä–æ–¥–∞–º–∏: {row.get('Users With Cities', 'N/A')}, "
                    f"–ë–µ–∑ –≥–æ—Ä–æ–¥–æ–≤: {row.get('Users Without Cities', 'N/A')}, "
                    f"–ó–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω–æ: {row.get('Blocked Users', 'N/A')}, "
                    f"–û—à–∏–±–æ–∫: {row.get('Errors', 'N/A')}"
                    for _, row in last_updates.iterrows()
                ])
            else:
                last_updates_str = "–ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö"
            
            # –§–æ—Ä–º–∏—Ä—É–µ–º –æ—Ç—á–µ—Ç
            report = f"–û—Ç—á–µ—Ç –æ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π (–Ω–∞ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')})\n\n"
            report += f"–í—Å–µ–≥–æ –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: {len(all_users)}\n"
            report += f"–ê–∫—Ç–∏–≤–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: {active_users}\n"
            report += f"–û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –Ω–∞–∂–∞—Ç–∏–π –Ω–∞ –∫–Ω–æ–ø–∫—É '–û–±–Ω–æ–≤–∏—Ç—å': {total_refresh_clicks}\n"
            report += f"–û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –Ω–∞–∂–∞—Ç–∏–π –Ω–∞ –∫–Ω–æ–ø–∫–∏ –≥–æ—Ä–æ–¥–æ–≤: {total_city_clicks}\n"
            report += f"–û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–π –∫–æ–º–∞–Ω–¥—ã /start: {total_start_cmds}\n"
            report += f"–û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏—Ö –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π: {total_auto_updates}\n\n"
            
            report += "–ü–æ—Å–ª–µ–¥–Ω–∏–µ 5 –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏—Ö –æ–±–Ω–æ–≤–ª–µ–Ω–∏–π:\n" + last_updates_str + "\n\n"
            
            if not top_refresh_users.empty:
                report += "–¢–æ–ø-5 –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –ø–æ –Ω–∞–∂–∞—Ç–∏—è–º –∫–Ω–æ–ø–∫–∏ '–û–±–Ω–æ–≤–∏—Ç—å':\n"
                for idx, row in top_refresh_users.iterrows():
                    report += f"- –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å {row['User ID']}: {row['Count']} –Ω–∞–∂–∞—Ç–∏–π\n"
                report += "\n"
            
            if not top_start_users.empty:
                report += "–¢–æ–ø-5 –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –ø–æ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—é –∫–æ–º–∞–Ω–¥—ã /start:\n"
                for idx, row in top_start_users.iterrows():
                    report += f"- –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å {row['User ID']}: {row['Count']} –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–π\n"
                report += "\n"
            
            if not top_cities.empty:
                report += "–¢–æ–ø-5 —Å–∞–º—ã—Ö –ø–æ–ø—É–ª—è—Ä–Ω—ã—Ö –≥–æ—Ä–æ–¥–æ–≤:\n"
                for idx, row in top_cities.iterrows():
                    report += f"- {row['City']}: {row['Count']} –Ω–∞–∂–∞—Ç–∏–π\n"
            
            # –°–æ–∑–¥–∞–µ–º –æ—Ç—á–µ—Ç –≤ Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # –î–æ–±–∞–≤–ª—è–µ–º –≤–∫–ª–∞–¥–∫—É —Å –æ—Ç—á–µ—Ç–æ–º
                workbook = writer.book
                if 'Activity Report' in workbook.sheetnames:
                    workbook.remove(workbook['Activity Report'])
                
                worksheet = workbook.create_sheet('Activity Report')
                worksheet['A1'] = report
                worksheet['A1'].font = Font(size=12)
                worksheet['A1'].alignment = Alignment(wrap_text=True)
                
                # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º —à–∏—Ä–∏–Ω—É –∫–æ–ª–æ–Ω–∫–∏
                worksheet.column_dimensions['A'].width = 100
                
            activity_logger.info(f"–°–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω –æ—Ç—á–µ—Ç –æ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π")
            return report
    except Exception as e:
        activity_logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –æ—Ç—á–µ—Ç–∞ –æ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏: {e}")
        return f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –æ—Ç—á–µ—Ç–∞: {e}"

def load_activity_data():
    """–ó–∞–≥—Ä—É–∂–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –∏–∑ Excel –≤ –ø–∞–º—è—Ç—å"""
    try:
        if os.path.exists(ACTIVITY_LOG_FILE):
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∫–Ω–æ–ø–∫–∏ "–û–±–Ω–æ–≤–∏—Ç—å"
            df_refresh = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Refresh Button')
            for idx, row in df_refresh.iterrows():
                activity_data['refresh'][row['User ID']] = row['Count']
            
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∫–Ω–æ–ø–æ–∫ –≥–æ—Ä–æ–¥–æ–≤
            df_cities = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='City Buttons')
            for idx, row in df_cities.iterrows():
                user_id = row['User ID']
                city = row['City']
                count = row['Count']
                
                if user_id not in activity_data['cities']:
                    activity_data['cities'][user_id] = {}
                
                activity_data['cities'][user_id][city] = count
            
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∫–æ–º–∞–Ω–¥—ã /start
            try:
                df_start = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Start Command')
                for idx, row in df_start.iterrows():
                    activity_data['start'][row['User ID']] = row['Count']
            except Exception as e:
                activity_logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –æ –∫–æ–º–∞–Ω–¥–µ /start: {e}")
            
            # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏–π
            try:
                df_auto_updates = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Auto Updates')
                if not df_auto_updates.empty:
                    if 'Messages Sent' in df_auto_updates.columns:
                        activity_data['auto_updates']['total_sent'] = df_auto_updates['Messages Sent'].sum()
                    elif 'Users Count' in df_auto_updates.columns:
                        activity_data['auto_updates']['total_sent'] = df_auto_updates['Users Count'].sum()
                    
                    last_update = df_auto_updates.iloc[-1]
                    activity_data['auto_updates']['last_update'] = f"{last_update['Date']} {last_update['Time']}"
            except Exception as e:
                activity_logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –æ–± –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏—è—Ö: {e}")
                
            activity_logger.info(f"–ó–∞–≥—Ä—É–∂–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –∏–∑ {ACTIVITY_LOG_FILE}")
    except Exception as e:
        activity_logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–≥—Ä—É–∑–∫–µ –¥–∞–Ω–Ω—ã—Ö –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏: {e}")

# –û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã /language –¥–ª—è –≤—ã–±–æ—Ä–∞ —è–∑—ã–∫–∞
@bot.message_handler(commands=['language'])
def language_command(message):
    try:
        chat_id = message.chat.id
        user_languages = load_user_languages()
        language = get_user_language(chat_id, user_languages)

        markup = types.InlineKeyboardMarkup(row_width=2)
        buttons = [
            types.InlineKeyboardButton("üá∑üá∫ –†—É—Å—Å–∫–∏–π", callback_data="lang_ru"),
            types.InlineKeyboardButton("üá¨üáß English", callback_data="lang_en"),
            types.InlineKeyboardButton("üá™üá∏ Espa√±ol", callback_data="lang_es"),
            types.InlineKeyboardButton("üá©üá™ Deutsch", callback_data="lang_de")
        ]
        markup.add(*buttons)

        selection_text = t('language_selection', language)
        bot.send_message(chat_id, selection_text, reply_markup=markup)

    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –≤ –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–µ /language: {e}")
        bot.send_message(message.chat.id, "Error / –û—à–∏–±–∫–∞")

# –û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã /start
@bot.message_handler(commands=['start'])
def start(message):
    try:
        chat_id = message.chat.id

        # –î–æ–±–∞–≤–ª—è–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ –æ–±—â–∏–π —Å–ø–∏—Å–æ–∫
        add_user_to_all_users(chat_id)

        # –û—Ç—Å–ª–µ–∂–∏–≤–∞–µ–º –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ –∫–æ–º–∞–Ω–¥—ã /start
        update_start_log(chat_id)

        if chat_id in last_messages:
            del last_messages[chat_id]

        user_data = load_user_data()
        weather_cache = load_weather_cache()
        user_languages = load_user_languages()
        send_welcome_message(chat_id, user_data, weather_cache, user_languages)

    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –≤ –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–µ /start: {e}")
        user_languages = load_user_languages()
        language = get_user_language(message.chat.id, user_languages)
        bot.send_message(message.chat.id, t('unknown_command', language))
        user_data = load_user_data()
        weather_cache = load_weather_cache()
        send_welcome_message(message.chat.id, user_data, weather_cache, user_languages)

# –û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã /stats –¥–ª—è –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –æ—Ç—á–µ—Ç–∞ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏
@bot.message_handler(commands=['stats'])
def stats(message):
    try:
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –∑–∞–ø—Ä–æ—Å –æ—Ç –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞ (–º–æ–∂–Ω–æ –Ω–∞—Å—Ç—Ä–æ–∏—Ç—å —Å–ø–∏—Å–æ–∫ –∞–¥–º–∏–Ω–æ–≤)
        admin_ids = [message.chat.id]  # –ó–¥–µ—Å—å –º–æ–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å —Å–ø–∏—Å–æ–∫ ID –∞–¥–º–∏–Ω–æ–≤
        
        if message.chat.id in admin_ids:
            report = generate_activity_report()
            bot.send_message(message.chat.id, f"–û—Ç—á–µ—Ç –æ–± –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π:\n\n{report}")
        else:
            bot.send_message(message.chat.id, "–£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ —ç—Ç–æ–π –∫–æ–º–∞–Ω–¥–µ.")
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –≤ –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–µ /stats: {e}")
        bot.send_message(message.chat.id, "–ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –æ—Ç—á–µ—Ç–∞.")

# –û–±—Ä–∞–±–æ—Ç—á–∏–∫ –∫–æ–º–∞–Ω–¥—ã /check_users –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ —Å—Ç–∞—Ç—É—Å–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
@bot.message_handler(commands=['check_users'])
def check_users(message):
    try:
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –∑–∞–ø—Ä–æ—Å –æ—Ç –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞
        admin_ids = [message.chat.id]  # –ó–¥–µ—Å—å –º–æ–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å —Å–ø–∏—Å–æ–∫ ID –∞–¥–º–∏–Ω–æ–≤
        
        if message.chat.id in admin_ids:
            all_users = load_all_users()
            user_data = load_user_data()
            
            active_users = 0
            inactive_users = 0
            users_with_cities = 0
            users_without_cities = 0
            additional_list_users = 0
            start_command_users = 0
            
            for user_id_str, user_info in all_users.items():
                if user_info.get('active', True):
                    active_users += 1
                    user_cities = get_user_cities(int(user_id_str), user_data)
                    
                    if user_cities:
                        users_with_cities += 1
                    else:
                        users_without_cities += 1
                else:
                    inactive_users += 1
                
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∏—Å—Ç–æ—á–Ω–∏–∫ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
                source = user_info.get('source', 'unknown')
                if source == 'additional_list':
                    additional_list_users += 1
                elif source == 'start_command':
                    start_command_users += 1
            
            check_report = f"–ü—Ä–æ–≤–µ—Ä–∫–∞ —Å—Ç–∞—Ç—É—Å–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π (–Ω–∞ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}):\n\n"
            check_report += f"üìä –û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞:\n"
            check_report += f"‚Ä¢ –í—Å–µ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: {len(all_users)}\n"
            check_report += f"‚Ä¢ –ê–∫—Ç–∏–≤–Ω—ã—Ö: {active_users}\n"
            check_report += f"‚Ä¢ –ù–µ–∞–∫—Ç–∏–≤–Ω—ã—Ö (–∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–ª–∏ –±–æ—Ç–∞): {inactive_users}\n\n"
            
            check_report += f"üèôÔ∏è –ü–æ –Ω–∞–ª–∏—á–∏—é –≥–æ—Ä–æ–¥–æ–≤:\n"
            check_report += f"‚Ä¢ –° –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã–º–∏ –≥–æ—Ä–æ–¥–∞–º–∏: {users_with_cities}\n"
            check_report += f"‚Ä¢ –ë–µ–∑ –≥–æ—Ä–æ–¥–æ–≤: {users_without_cities}\n\n"
            
            check_report += f"üìù –ü–æ –∏—Å—Ç–æ—á–Ω–∏–∫–∞–º:\n"
            check_report += f"‚Ä¢ –ò–∑ –¥–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ–≥–æ —Å–ø–∏—Å–∫–∞: {additional_list_users}\n"
            check_report += f"‚Ä¢ –ß–µ—Ä–µ–∑ –∫–æ–º–∞–Ω–¥—É /start: {start_command_users}\n\n"
            
            success_rate = (active_users / len(all_users)) * 100 if len(all_users) > 0 else 0
            check_report += f"‚úÖ –ü—Ä–æ—Ü–µ–Ω—Ç –∞–∫—Ç–∏–≤–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: {success_rate:.1f}%\n"
            
            bot.send_message(message.chat.id, check_report)
        else:
            bot.send_message(message.chat.id, "–£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ —ç—Ç–æ–π –∫–æ–º–∞–Ω–¥–µ.")
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –≤ –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–µ /check_users: {e}")
        bot.send_message(message.chat.id, "–ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π.")

def get_and_send_weather(chat_id, city, user_data, weather_cache, user_languages, message_id=None, force_new_message=False):
    language = get_user_language(chat_id, user_languages)

    # –ò—Å–ø–æ–ª—å–∑—É–µ–º —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤—É—é—â–∏–π —è–∑—ã–∫ –¥–ª—è OpenWeather API
    api_lang_map = {'ru': 'ru', 'en': 'en', 'es': 'es', 'de': 'de'}
    api_lang = api_lang_map.get(language, 'en')

    url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&units=metric&lang={api_lang}&appid={OPENWEATHER_API_KEY}'
    
    try:
        weather_data = requests.get(url).json()
        
        if weather_data.get('cod') == '404':
            error_text = t('city_not_found', language)
            if force_new_message and message_id:
                try:
                    bot.delete_message(chat_id, message_id)
                except Exception as e:
                    logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å –ø—Ä–µ–¥—ã–¥—É—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ: {e}")

                error_msg = send_new_message(chat_id, error_text)
            else:
                msg_info = last_messages.get(chat_id, {})
                msg_id = message_id if message_id else msg_info.get('message_id')
                error_msg = update_message(chat_id, msg_id, error_text)
            
            if error_msg and hasattr(error_msg, 'message_id'):
                last_messages[chat_id] = {
                    'message_id': error_msg.message_id,
                    'timestamp': time.time()
                }
            return False
            
        temperature = round(weather_data['main']['temp'], 2)
        temperature_feels = round(weather_data['main']['feels_like'], 2)
        weather_description = weather_data['weather'][0]['description']
        wind_speed = weather_data['wind']['speed']
        timezone_offset = weather_data.get('timezone', 0)  # –°–º–µ—â–µ–Ω–∏–µ —á–∞—Å–æ–≤–æ–≥–æ –ø–æ—è—Å–∞ –≤ —Å–µ–∫—É–Ω–¥–∞—Ö

        # –û–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ —Ç–µ–∫—É—â–µ–≥–æ –≤—Ä–µ–º–µ–Ω–∏ –≥–æ–¥–∞
        from datetime import datetime
        current_date = datetime.now()

        month = current_date.month
        if month in [12, 1, 2]:
            season = '–∑–∏–º–∞'
        elif month in [3, 4, 5]:
            season = '–≤–µ—Å–Ω–∞'
        elif month in [6, 7, 8]:
            season = '–ª–µ—Ç–æ'
        else:  # 9, 10, 11
            season = '–æ—Å–µ–Ω—å'

        weather_emoji = get_weather_emoji(weather_description)
        # –ò—Å–ø–æ–ª—å–∑—É–µ–º –æ–±–Ω–æ–≤–ª–µ–Ω–Ω—É—é —Ñ—É–Ω–∫—Ü–∏—é —Å —É—á–µ—Ç–æ–º –≤—Ä–µ–º–µ–Ω–∏ –≥–æ–¥–∞, –≤—Ä–µ–º–µ–Ω–∏ —Å—É—Ç–æ–∫, —á–∞—Å–æ–≤–æ–≥–æ –ø–æ—è—Å–∞ –≥–æ—Ä–æ–¥–∞ –∏ —è–∑—ã–∫–∞
        clothes_advice = get_clothing_advice(temperature, weather_description, season, None, wind_speed, timezone_offset, language)

        temp_str = f"+{temperature}" if temperature > 0 else f"{temperature}"
        temp_feels_str = f"+{temperature_feels}" if temperature_feels > 0 else f"{temperature_feels}"
        weather_description_cap = weather_description[0].upper() + weather_description[1:]

        # –ü–æ–ª—É—á–∞–µ–º –ª–æ–∫–∞–ª—å–Ω–æ–µ –≤—Ä–µ–º—è –≥–æ—Ä–æ–¥–∞
        local_formatted_time = format_local_date_time(timezone_offset)

        weather_message = (
            f'{weather_emoji} {city} {weather_description_cap}\n'
            f'üå°Ô∏è {t("temperature", language)} {temp_str}¬∞C\n'
            f'üå°Ô∏è {t("feels_like", language)} {temp_feels_str}¬∞C\n'
            f'üí® {t("wind_speed", language)} | {wind_speed} –º/—Å\n'
            f'{clothes_advice}\n'
            f'‚è±Ô∏è {t("update_time", language)}: {local_formatted_time}'
        )
        
        cache_data = {
            'temp': temperature,
            'emoji': weather_emoji,
            'description': weather_description,
            'wind_speed': wind_speed,
            'updated_at': current_time
        }
        update_cached_weather(city, cache_data, weather_cache)
        
        markup = types.InlineKeyboardMarkup(row_width=2)
        user_cities = get_user_cities(chat_id, user_data)
        
        if city in user_cities:
            markup.add(
                types.InlineKeyboardButton("–£–¥–∞–ª–∏—Ç—å –≥–æ—Ä–æ–¥", callback_data=f"remove_{city}"),
                types.InlineKeyboardButton("–ù–∞–∑–∞–¥", callback_data="back")
            )
        else:
            markup.add(
                types.InlineKeyboardButton("–î–æ–±–∞–≤–∏—Ç—å –≥–æ—Ä–æ–¥", callback_data=f"add_{city}"),
                types.InlineKeyboardButton("–ù–∞–∑–∞–¥", callback_data="back")
            )
        
        if force_new_message:
            # –°–Ω–∞—á–∞–ª–∞ —É–¥–∞–ª—è–µ–º –ø—Ä–µ–¥—ã–¥—É—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
            try:
                if message_id:
                    bot.delete_message(chat_id, message_id)
            except Exception as e:
                logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å –ø—Ä–µ–¥—ã–¥—É—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ: {e}")
            
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –Ω–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
            sent_message = send_new_message(chat_id, weather_message, markup)
        else:
            # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
            msg_info = last_messages.get(chat_id, {})
            msg_id = message_id if message_id else msg_info.get('message_id')
            sent_message = update_message(chat_id, msg_id, weather_message, markup)
        
        if sent_message and hasattr(sent_message, 'message_id'):
            last_messages[chat_id] = {
                'message_id': sent_message.message_id,
                'timestamp': time.time()
            }
            
        return True
    
    except Exception as e:
        logger.error(f'–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ –ø–æ–≥–æ–¥—ã: {e}')
        
        error_text = t('weather_error', language)
        if force_new_message and message_id:
            try:
                bot.delete_message(chat_id, message_id)
            except Exception as e:
                logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å –ø—Ä–µ–¥—ã–¥—É—â–µ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ: {e}")

            error_msg = send_new_message(chat_id, error_text)
        else:
            msg_info = last_messages.get(chat_id, {})
            msg_id = message_id if message_id else msg_info.get('message_id')

            error_msg = update_message(chat_id, msg_id, error_text)
        
        if error_msg and hasattr(error_msg, 'message_id'):
            last_messages[chat_id] = {
                'message_id': error_msg.message_id,
                'timestamp': time.time()
            }
            
        return False

@bot.callback_query_handler(func=lambda call: True)
def callback_handler(call):
    user_data = load_user_data()
    weather_cache = load_weather_cache()
    user_languages = load_user_languages()

    try:
        # –û–±—Ä–∞–±–æ—Ç–∫–∞ –≤—ã–±–æ—Ä–∞ —è–∑—ã–∫–∞
        if call.data.startswith("lang_"):
            lang_code = call.data[5:]
            if lang_code in SUPPORTED_LANGUAGES:
                user_languages = set_user_language(call.message.chat.id, lang_code, user_languages)
                save_user_languages(user_languages)

                language = lang_code
                bot.edit_message_text(
                    t('language_changed', language),
                    call.message.chat.id,
                    call.message.message_id
                )

                # –û–±–Ω–æ–≤–ª—è–µ–º –ø—Ä–∏–≤–µ—Ç—Å—Ç–≤–µ–Ω–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å –Ω–æ–≤—ã–º —è–∑—ã–∫–æ–º
                time.sleep(1)
                send_welcome_message(call.message.chat.id, user_data, weather_cache, user_languages)

        elif call.data == "back":
            send_welcome_message(call.message.chat.id, user_data, weather_cache, user_languages, call.message.message_id)

        elif call.data == "refresh":
            # –û—Ç—Å–ª–µ–∂–∏–≤–∞–µ–º –Ω–∞–∂–∞—Ç–∏–µ –∫–Ω–æ–ø–∫–∏ "–û–±–Ω–æ–≤–∏—Ç—å"
            update_refresh_log(call.message.chat.id)

            # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –æ –ø–æ–≥–æ–¥–µ –¥–ª—è –≤—Å–µ—Ö –≥–æ—Ä–æ–¥–æ–≤ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
            user_cities = get_user_cities(call.message.chat.id, user_data)
            for city in user_cities:
                get_weather_data(city, weather_cache)
                time.sleep(0.5)  # –ù–µ–±–æ–ª—å—à–∞—è –∑–∞–¥–µ—Ä–∂–∫–∞ –º–µ–∂–¥—É –∑–∞–ø—Ä–æ—Å–∞–º–∏

            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –Ω–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ, –Ω–æ —Å–Ω–∞—á–∞–ª–∞ —É–¥–∞–ª—è–µ–º –ø—Ä–µ–¥—ã–¥—É—â–µ–µ
            send_welcome_message(call.message.chat.id, user_data, weather_cache, user_languages, call.message.message_id, force_new_message=True)

        elif call.data.startswith("add_"):
            city = call.data[4:]
            if add_user_city(call.message.chat.id, city, user_data):
                send_welcome_message(call.message.chat.id, user_data, weather_cache, user_languages, call.message.message_id)

        elif call.data.startswith("remove_"):
            city = call.data[7:]
            if remove_user_city(call.message.chat.id, city, user_data):
                send_welcome_message(call.message.chat.id, user_data, weather_cache, user_languages, call.message.message_id)

        elif call.data.startswith("city_"):
            city = call.data[5:]
            # –û—Ç—Å–ª–µ–∂–∏–≤–∞–µ–º –Ω–∞–∂–∞—Ç–∏–µ –∫–Ω–æ–ø–∫–∏ –≥–æ—Ä–æ–¥–∞
            update_city_log(call.message.chat.id, city)

            get_and_send_weather(call.message.chat.id, city, user_data, weather_cache, user_languages, call.message.message_id, force_new_message=False)

        bot.answer_callback_query(call.id)
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –≤ callback_handler: {e}")
        try:
            bot.answer_callback_query(call.id)
        except:
            pass

@bot.message_handler(func=lambda message: True)
def handle_all_messages(message):
    try:
        if message.text.startswith('/'):
            user_languages = load_user_languages()
            language = get_user_language(message.chat.id, user_languages)
            bot.send_message(message.chat.id, t('unknown_command', language))
        else:
            user_data = load_user_data()
            weather_cache = load_weather_cache()
            user_languages = load_user_languages()
            get_and_send_weather(message.chat.id, message.text, user_data, weather_cache, user_languages, last_messages.get(message.chat.id, {}).get('message_id'), force_new_message=False)
        
        try:
            bot.delete_message(message.chat.id, message.message_id)
        except Exception as e:
            logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å —Å–æ–æ–±—â–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è: {e}")
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Å–æ–æ–±—â–µ–Ω–∏—è: {e}")
        bot.send_message(message.chat.id, "–ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.")

@bot.inline_handler(func=lambda query: True)
def handle_inline_query(query):
    try:
        city = query.query.strip()
        
        if not city:
            # –ï—Å–ª–∏ –≥–æ—Ä–æ–¥ –Ω–µ —É–∫–∞–∑–∞–Ω, –ø—Ä–µ–¥–ª–∞–≥–∞–µ–º –≤–∞—Ä–∏–∞–Ω—Ç—ã —Å –∞–∫—Ç—É–∞–ª—å–Ω–æ–π –ø–æ–≥–æ–¥–æ–π
            # –ü–æ–ª—É—á–∞–µ–º –¥–∞–Ω–Ω—ã–µ –æ –ø–æ–≥–æ–¥–µ –¥–ª—è –ú–æ—Å–∫–≤—ã
            moscow_weather = get_weather_data("–ú–æ—Å–∫–≤–∞", load_weather_cache())
            if not moscow_weather:
                moscow_temp = "+14"
                moscow_temp_feels = "+12"
                moscow_description = "–û–±–ª–∞—á–Ω–æ —Å –ø—Ä–æ—è—Å–Ω–µ–Ω–∏—è–º–∏"
                moscow_wind = 6
                moscow_emoji = "‚õÖ"
            else:
                moscow_temp = f"+{moscow_weather['temp']}" if moscow_weather['temp'] > 0 else f"{moscow_weather['temp']}"
                moscow_temp_feels = f"+{moscow_weather['temp'] - 2}" if moscow_weather['temp'] > 0 else f"{moscow_weather['temp'] - 2}"
                moscow_description = "–û–±–ª–∞—á–Ω–æ —Å –ø—Ä–æ—è—Å–Ω–µ–Ω–∏—è–º–∏"
                moscow_wind = moscow_weather['wind_speed']
                moscow_emoji = moscow_weather['emoji']
            
            # –ü–æ–ª—É—á–∞–µ–º –¥–∞–Ω–Ω—ã–µ –æ –ø–æ–≥–æ–¥–µ –¥–ª—è –°–∞–Ω–∫—Ç-–ü–µ—Ç–µ—Ä–±—É—Ä–≥–∞
            spb_weather = get_weather_data("–°–∞–Ω–∫—Ç-–ü–µ—Ç–µ—Ä–±—É—Ä–≥", load_weather_cache())
            if not spb_weather:
                spb_temp = "+8"
                spb_temp_feels = "+3"
                spb_description = "–û–±–ª–∞—á–Ω–æ"
                spb_wind = 5
                spb_emoji = "‚òÅÔ∏è"
            else:
                spb_temp = f"+{spb_weather['temp']}" if spb_weather['temp'] > 0 else f"{spb_weather['temp']}"
                spb_temp_feels = f"+{spb_weather['temp'] - 5}" if spb_weather['temp'] > 0 else f"{spb_weather['temp'] - 5}"
                spb_description = "–û–±–ª–∞—á–Ω–æ"
                spb_wind = spb_weather['wind_speed']
                spb_emoji = spb_weather['emoji']
            
            # –û–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ —Ç–µ–∫—É—â–µ–≥–æ –≤—Ä–µ–º–µ–Ω–∏ –≥–æ–¥–∞
            from datetime import datetime
            current_date = datetime.now()

            month = current_date.month
            if month in [12, 1, 2]:
                season = '–∑–∏–º–∞'
            elif month in [3, 4, 5]:
                season = '–≤–µ—Å–Ω–∞'
            elif month in [6, 7, 8]:
                season = '–ª–µ—Ç–æ'
            else:  # 9, 10, 11
                season = '–æ—Å–µ–Ω—å'

            # –ü–æ–ª—É—á–∞–µ–º timezone –¥–ª—è –≥–æ—Ä–æ–¥–æ–≤
            moscow_timezone = moscow_weather.get('timezone', 10800) if moscow_weather else 10800  # –ú–æ—Å–∫–≤–∞ UTC+3
            spb_timezone = spb_weather.get('timezone', 10800) if spb_weather else 10800  # –°–ü–ë UTC+3

            # –§–æ—Ä–º–∏—Ä—É–µ–º —Å–æ–≤–µ—Ç—ã –ø–æ –æ–¥–µ–∂–¥–µ —Å —É—á–µ—Ç–æ–º –≤—Ä–µ–º–µ–Ω–∏ –≥–æ–¥–∞ –∏ —á–∞—Å–æ–≤–æ–≥–æ –ø–æ—è—Å–∞ –≥–æ—Ä–æ–¥–∞
            moscow_clothes_advice = get_clothing_advice(float(moscow_temp.replace("+", "")), moscow_description, season, None, moscow_wind, moscow_timezone)
            spb_clothes_advice = get_clothing_advice(float(spb_temp.replace("+", "")), spb_description, season, None, spb_wind, spb_timezone)

            # –ü–æ–ª—É—á–∞–µ–º –ª–æ–∫–∞–ª—å–Ω–æ–µ –≤—Ä–µ–º—è –¥–ª—è –∫–∞–∂–¥–æ–≥–æ –≥–æ—Ä–æ–¥–∞
            moscow_local_time = format_local_date_time(moscow_timezone)
            spb_local_time = format_local_date_time(spb_timezone)

            results = [
                types.InlineQueryResultArticle(
                    id="1",
                    title="–ú–æ—Å–∫–≤–∞",
                    description=f"{moscow_temp}¬∞C, {moscow_description}",
                    input_message_content=types.InputTextMessageContent(
                        message_text=f"{moscow_emoji} –ú–æ—Å–∫–≤–∞ {moscow_description}\nüå°Ô∏è t¬∞ {moscow_temp}¬∞C\nüå°Ô∏è t¬∞–æ—â—É—â. {moscow_temp_feels}¬∞C\nüí® –°–∫–æ—Ä–æ—Å—Ç—å –≤–µ—Ç—Ä–∞ | {moscow_wind} –º/—Å\n {moscow_clothes_advice}\n‚è±Ô∏è –í—Ä–µ–º—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è: {moscow_local_time}"
                    )
                ),
                types.InlineQueryResultArticle(
                    id="2",
                    title="–°–∞–Ω–∫—Ç-–ü–µ—Ç–µ—Ä–±—É—Ä–≥",
                    description=f"{spb_temp}¬∞C, {spb_description}",
                    input_message_content=types.InputTextMessageContent(
                        message_text=f"{spb_emoji} –°–∞–Ω–∫—Ç-–ü–µ—Ç–µ—Ä–±—É—Ä–≥ {spb_description}\nüå°Ô∏è t¬∞ {spb_temp}¬∞C\nüå°Ô∏è t¬∞–æ—â—É—â. {spb_temp_feels}¬∞C\nüí® –°–∫–æ—Ä–æ—Å—Ç—å –≤–µ—Ç—Ä–∞ | {spb_wind} –º/—Å\n {spb_clothes_advice}\n‚è±Ô∏è –í—Ä–µ–º—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è: {spb_local_time}"
                    )
                )
            ]
            bot.answer_inline_query(query.id, results)
            return

        # –ü–æ–ª—É—á–∞–µ–º —Ä–µ–∞–ª—å–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ –æ –ø–æ–≥–æ–¥–µ
        url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&units=metric&lang=ru&appid={OPENWEATHER_API_KEY}'
        response = requests.get(url)
        response.raise_for_status()
        weather_data = response.json()
        
        if weather_data.get('cod') == '404':
            bot.answer_inline_query(query.id, [])
            return
            
        temperature = round(weather_data['main']['temp'], 2)
        temperature_feels = round(weather_data['main']['feels_like'], 2)
        weather_description = weather_data['weather'][0]['description']
        wind_speed = weather_data['wind']['speed']
        timezone_offset = weather_data.get('timezone', 0)  # –°–º–µ—â–µ–Ω–∏–µ —á–∞—Å–æ–≤–æ–≥–æ –ø–æ—è—Å–∞ –≤ —Å–µ–∫—É–Ω–¥–∞—Ö

        # –û–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ —Ç–µ–∫—É—â–µ–≥–æ –≤—Ä–µ–º–µ–Ω–∏ –≥–æ–¥–∞
        from datetime import datetime
        current_date = datetime.now()

        month = current_date.month
        if month in [12, 1, 2]:
            season = '–∑–∏–º–∞'
        elif month in [3, 4, 5]:
            season = '–≤–µ—Å–Ω–∞'
        elif month in [6, 7, 8]:
            season = '–ª–µ—Ç–æ'
        else:  # 9, 10, 11
            season = '–æ—Å–µ–Ω—å'

        weather_emoji = get_weather_emoji(weather_description)
        clothes_advice = get_clothing_advice(temperature, weather_description, season, None, wind_speed, timezone_offset)

        temp_str = f"+{temperature}" if temperature > 0 else f"{temperature}"
        temp_feels_str = f"+{temperature_feels}" if temperature_feels > 0 else f"{temperature_feels}"
        weather_description_cap = weather_description[0].upper() + weather_description[1:]

        # –ü–æ–ª—É—á–∞–µ–º –ª–æ–∫–∞–ª—å–Ω–æ–µ –≤—Ä–µ–º—è –≥–æ—Ä–æ–¥–∞
        local_formatted_time = format_local_date_time(timezone_offset)

        # –§–æ—Ä–º–∏—Ä—É–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ –≤ —Ç–æ—á–Ω–æ–º —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤–∏–∏ —Å –ø—Ä–∏–º–µ—Ä–æ–º
        message_text = (
            f"{weather_emoji} {city} {weather_description_cap}\n"
            f"üå°Ô∏è t¬∞ {temp_str}¬∞C\n"
            f"üå°Ô∏è t¬∞–æ—â—É—â. {temp_feels_str}¬∞C\n"
            f"üí® –°–∫–æ—Ä–æ—Å—Ç—å –≤–µ—Ç—Ä–∞ | {wind_speed} –º/—Å\n"
            f"{clothes_advice}\n"
            f"‚è±Ô∏è –í—Ä–µ–º—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è: {local_formatted_time}"
        )

        result = types.InlineQueryResultArticle(
            id="1",
            title=f"–ü–æ–≥–æ–¥–∞ –≤ {city}",
            description=f"{temp_str}¬∞C, {weather_description}",
            input_message_content=types.InputTextMessageContent(
                message_text=message_text
            )
        )

        bot.answer_inline_query(query.id, [result])

    except Exception as e:
        logger.error(f"Inline error: {e}")
        bot.answer_inline_query(query.id, [])

def clean_message_cache():
    while True:
        try:
            current_time = time.time()
            to_delete = [chat_id for chat_id, msg_info in last_messages.items() 
                        if current_time - msg_info.get('timestamp', 0) > 604800]
            
            for chat_id in to_delete:
                del last_messages[chat_id]
            
            time.sleep(86400)
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –≤ clean_message_cache: {e}")
            time.sleep(3600)

def update_all_weather_info():
    while True:
        try:
            logger.info(f"–û–±–Ω–æ–≤–ª–µ–Ω–∏–µ –∫—ç—à–∞ –ø–æ–≥–æ–¥—ã... {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            user_data = load_user_data()
            weather_cache = load_weather_cache()
            update_weather_cache(weather_cache, user_data)
            
            logger.info(f"–û–±–Ω–æ–≤–ª–µ–Ω–∏–µ –∫—ç—à–∞ –ø–æ–≥–æ–¥—ã –∑–∞–≤–µ—Ä—à–µ–Ω–æ! {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            time.sleep(600)
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±–Ω–æ–≤–ª–µ–Ω–∏–∏ –∫—ç—à–∞: {e}")
            time.sleep(60)

if __name__ == '__main__':
    if not os.path.exists(USER_DATA_FILE):
        save_user_data({})

    if not os.path.exists(WEATHER_CACHE_FILE):
        save_weather_cache({})

    if not os.path.exists(ALL_USERS_FILE):
        save_all_users({})

    if not os.path.exists(USER_LANGUAGES_FILE):
        save_user_languages({})
    
    # –î–æ–±–∞–≤–ª—è–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–∑ —Å–ø–∏—Å–∫–∞ –≤ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö
    added_users = add_additional_users_to_all_users()
    
    # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º Excel-—Ñ–∞–π–ª –¥–ª—è –æ—Ç—Å–ª–µ–∂–∏–≤–∞–Ω–∏—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏
    initialize_excel_file()
    
    # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏
    load_activity_data()
    
    # –ó–∞–ø—É—Å–∫–∞–µ–º –ø–æ—Ç–æ–∫–∏ –¥–ª—è —Ñ–æ–Ω–æ–≤—ã—Ö –∑–∞–¥–∞—á
    threading.Thread(target=clean_message_cache, daemon=True).start()
    threading.Thread(target=update_all_weather_info, daemon=True).start()
    threading.Thread(target=auto_update_users, daemon=True).start()
    
    logger.info("–ë–æ—Ç GidMeteo –∑–∞–ø—É—â–µ–Ω")
    print("–ë–æ—Ç GidMeteo –∑–∞–ø—É—â–µ–Ω. –ù–∞–∂–º–∏—Ç–µ Ctrl+C –¥–ª—è –æ—Å—Ç–∞–Ω–æ–≤–∫–∏.")
    print("–û—Ç—Å–ª–µ–∂–∏–≤–∞–Ω–∏–µ –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏ –≤–∫–ª—é—á–µ–Ω–æ. –î–∞–Ω–Ω—ã–µ —Å–æ—Ö—Ä–∞–Ω—è—é—Ç—Å—è –≤ —Ñ–∞–π–ª:", ACTIVITY_LOG_FILE)
    print("–ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏–µ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –±—É–¥—É—Ç –æ—Ç–ø—Ä–∞–≤–ª—è—Ç—å—Å—è –∫–∞–∂–¥—ã–µ 4 —á–∞—Å–∞ (00:01, 04:01, 08:01, 12:01, 16:01, 20:01)")
    print(f"–î–æ–±–∞–≤–ª–µ–Ω–æ {added_users} –Ω–æ–≤—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–∑ –¥–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ–≥–æ —Å–ø–∏—Å–∫–∞")
    print("–î–æ—Å—Ç—É–ø–Ω—ã–µ –∫–æ–º–∞–Ω–¥—ã –¥–ª—è –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞:")
    print("  /stats - –ø–æ–¥—Ä–æ–±–Ω—ã–π –æ—Ç—á–µ—Ç –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç–∏")
    print("  /check_users - –ø—Ä–æ–≤–µ—Ä–∫–∞ —Å—Ç–∞—Ç—É—Å–∞ –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π")
    
    try:
        while True:
            try:
                bot.polling(none_stop=True, interval=0, timeout=60)
            except Exception as e:
                logger.error(f'–°—Ä–∞–±–æ—Ç–∞–ª–æ –∏—Å–∫–ª—é—á–µ–Ω–∏–µ: {e}')
                time.sleep(10)
    except KeyboardInterrupt:
        logger.info("–ü–æ–ª—É—á–µ–Ω —Å–∏–≥–Ω–∞–ª Ctrl+C. –û—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞—é –±–æ—Ç–∞...")
        print("\n–ë–æ—Ç –æ—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω.")