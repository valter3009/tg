"""
Telegram бот GidMeteo для получения информации о погоде
и персонализированных советов по одежде
"""

import telebot
import requests
from telebot import types
import json
import os
import time
import threading
from datetime import datetime, time as dt_time
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, Alignment
from clothes_advice import get_clothing_advice
import config
from flask import Flask
from threading import Thread

# Настройка логирования
logging.basicConfig(
    filename='bot_errors.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Настройка логирования для активности
activity_logger = logging.getLogger('activity_logger')
activity_handler = logging.FileHandler('activity_tracker.log')
activity_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
activity_logger.addHandler(activity_handler)
activity_logger.setLevel(logging.INFO)

bot = telebot.TeleBot(config.TELEGRAM_TOKEN)

USER_DATA_FILE = 'user_cities.json'
WEATHER_CACHE_FILE = 'weather_cache.json'
ALL_USERS_FILE = 'all_users.json'  # Новый файл для хранения всех пользователей
CACHE_UPDATE_INTERVAL = 3600
ACTIVITY_LOG_FILE = 'bot_activity_log.xlsx'
AUTO_UPDATE_LOG_FILE = 'auto_updates.log'

# Список пользователей для рассылки
ADDITIONAL_USERS = [
    471789857, 425748474, 6118023060, 6488366997, 5934413419, 6615704791, 
    1134118381, 1823348752, 6579300547, 5174302370, 1344487460, 7791445179, 
    1276348447, 278283980, 6556640321, 1521820146, 7523695427, 7880850349, 
    832185475, 149653247, 1775572520, 7643533302, 352808232, 7456672724, 
    5969931672, 993675994, 543397394, 5935464436, 1812257315, 6260364812, 
    434939312
]

# Хранилище последних сообщений с timestamp
last_messages = {}

# Структура данных для хранения активности
activity_data = {
    'refresh': {},  # user_id: count
    'cities': {},   # user_id: {city: count}
    'start': {},     # user_id: count - добавлен счетчик для /start
    'auto_updates': {'total_sent': 0, 'last_update': None}  # Статистика автообновлений
}

# Мьютекс для работы с Excel-файлом активности
excel_mutex = threading.Lock()

# Функции для работы с пользовательскими данными
def load_user_data():
    try:
        if os.path.exists(USER_DATA_FILE):
            with open(USER_DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f"Ошибка загрузки user_data: {e}")
        return {}

def save_user_data(data):
    try:
        with open(USER_DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"Ошибка сохранения user_data: {e}")

# Новые функции для работы со всеми пользователями
def load_all_users():
    try:
        if os.path.exists(ALL_USERS_FILE):
            with open(ALL_USERS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f"Ошибка загрузки all_users: {e}")
        return {}

def save_all_users(data):
    try:
        with open(ALL_USERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"Ошибка сохранения all_users: {e}")

def add_additional_users_to_all_users():
    """Добавляет дополнительных пользователей в список всех пользователей"""
    all_users = load_all_users()
    added_count = 0
    
    for user_id in ADDITIONAL_USERS:
        user_id_str = str(user_id)
        
        if user_id_str not in all_users:
            all_users[user_id_str] = {
                'first_start': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'active': True,
                'source': 'additional_list'  # Помечаем источник
            }
            added_count += 1
        else:
            # Обновляем статус активности для существующих пользователей
            all_users[user_id_str]['active'] = True
    
    if added_count > 0:
        save_all_users(all_users)
        logger.info(f"Добавлено {added_count} пользователей из дополнительного списка")
    
    return added_count

def add_user_to_all_users(user_id):
    """Добавляет пользователя в список всех пользователей"""
    all_users = load_all_users()
    user_id_str = str(user_id)
    
    if user_id_str not in all_users:
        all_users[user_id_str] = {
            'first_start': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'active': True,
            'source': 'start_command'  # Помечаем источник
        }
        save_all_users(all_users)
        return True
    else:
        # Обновляем статус активности
        all_users[user_id_str]['active'] = True
        save_all_users(all_users)
    return False

def deactivate_user(user_id):
    """Деактивирует пользователя (например, если он заблокировал бота)"""
    all_users = load_all_users()
    user_id_str = str(user_id)
    
    if user_id_str in all_users:
        all_users[user_id_str]['active'] = False
        save_all_users(all_users)

def load_weather_cache():
    try:
        if os.path.exists(WEATHER_CACHE_FILE):
            with open(WEATHER_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.error(f"Ошибка загрузки weather_cache: {e}")
        return {}

def save_weather_cache(data):
    try:
        with open(WEATHER_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logger.error(f"Ошибка сохранения weather_cache: {e}")

def get_cached_weather(city, weather_cache):
    return weather_cache.get(city.lower())

def update_cached_weather(city, weather_data, weather_cache):
    weather_cache[city.lower()] = weather_data
    save_weather_cache(weather_cache)

def get_user_cities(user_id, user_data):
    user_id_str = str(user_id)
    return user_data.get(user_id_str, [])

def add_user_city(user_id, city, user_data):
    user_id_str = str(user_id)
    
    if user_id_str not in user_data:
        user_data[user_id_str] = []
    
    if city not in user_data[user_id_str]:
        user_data[user_id_str].append(city)
        save_user_data(user_data)
        return True
    return False

def remove_user_city(user_id, city, user_data):
    user_id_str = str(user_id)
    
    if user_id_str in user_data and city in user_data[user_id_str]:
        user_data[user_id_str].remove(city)
        save_user_data(user_data)
        return True
    return False

def get_weather_emoji(description):
    description = description.lower()
    
    if any(word in description for word in ['ясно', 'чистое небо', 'безоблачно']):
        return '☀️'
    elif any(word in description for word in ['облачно с прояснениями', 'переменная облачность']):
        return '⛅'
    elif any(word in description for word in ['облачно', 'пасмурно']):
        return '☁️'
    elif any(word in description for word in ['дождь', 'ливень']):
        return '🌧️'
    elif any(word in description for word in ['гроза']):
        return '⛈️'
    elif any(word in description for word in ['снег', 'снегопад']):
        return '❄️'
    elif any(word in description for word in ['туман', 'мгла']):
        return '🌫️'
    else:
        return '🌦️'

def get_weather_data(city, weather_cache):
    url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&units=metric&lang=ru&appid={config.OPENWEATHER_API_KEY}'
    
    try:
        response = requests.get(url)
        response.raise_for_status()
        weather_data = response.json()
        
        if weather_data.get('cod') == '404':
            return None
            
        temperature = round(weather_data['main']['temp'], 1)
        weather_description = weather_data['weather'][0]['description']
        wind_speed = int(round(weather_data['wind']['speed']))
        
        weather_emoji = get_weather_emoji(weather_description)
        
        result = {
            'temp': temperature,
            'emoji': weather_emoji,
            'description': weather_description,
            'wind_speed': wind_speed,
            'updated_at': int(time.time())
        }
        
        update_cached_weather(city, result, weather_cache)
        return result
    
    except Exception as e:
        logger.error(f'Ошибка при получении погоды для {city}: {e}')
        return None

def format_date_time(timestamp):
    return datetime.fromtimestamp(timestamp).strftime('%d.%m.%Y %H:%M:%S')

def update_weather_cache(weather_cache, user_data):
    all_cities = set()
    
    for user_id, cities in user_data.items():
        for city in cities:
            all_cities.add(city)
    
    for city in all_cities:
        try:
            get_weather_data(city, weather_cache)
            time.sleep(1)
        except Exception as e:
            logger.error(f"Ошибка при обновлении погоды для {city}: {e}")

def create_cities_keyboard(user_id, user_data, weather_cache):
    markup = types.InlineKeyboardMarkup(row_width=1)
    user_cities = get_user_cities(user_id, user_data)
    
    for city in user_cities:
        cached_weather = get_cached_weather(city, weather_cache)
        current_time = int(time.time())
        
        if not cached_weather or (current_time - cached_weather.get('updated_at', 0)) > 3600:
            weather_data = get_weather_data(city, weather_cache)
            if weather_data:
                temp_str = f"+{weather_data['temp']}" if weather_data['temp'] > 0 else f"{weather_data['temp']}"
                button_text = f"{weather_data['emoji']} {city} {temp_str}°C 💨{weather_data['wind_speed']}м/с"
            else:
                button_text = city
        else:
            temp_str = f"+{cached_weather['temp']}" if cached_weather['temp'] > 0 else f"{cached_weather['temp']}"
            button_text = f"{cached_weather['emoji']} {city} {temp_str}°C 💨{cached_weather['wind_speed']}м/с"
            
        markup.add(types.InlineKeyboardButton(text=button_text, callback_data=f"city_{city}"))
    
    # Добавляем кнопку "Обновить", только если есть города
    if user_cities:
        markup.add(types.InlineKeyboardButton(text="🔄 Обновить", callback_data="refresh"))
    
    return markup

def send_new_message(chat_id, text, markup=None, parse_mode='Markdown'):
    try:
        sent_msg = bot.send_message(
            chat_id,
            text,
            reply_markup=markup,
            parse_mode=parse_mode
        )
        last_messages[chat_id] = {
            'message_id': sent_msg.message_id,
            'timestamp': time.time()
        }
        return sent_msg
    except Exception as e:
        logger.error(f"Ошибка при отправке нового сообщения: {e}")
        return None

def update_message(chat_id, message_id, text, markup=None, parse_mode='Markdown'):
    try:
        if message_id:
            try:
                return bot.edit_message_text(
                    text=text,
                    chat_id=chat_id,
                    message_id=message_id,
                    reply_markup=markup,
                    parse_mode=parse_mode
                )
            except telebot.apihelper.ApiTelegramException as e:
                if "message to edit not found" in str(e).lower():
                    logger.info(f"Сообщение {message_id} в чате {chat_id} не найдено, отправляем новое")
                    if chat_id in last_messages:
                        del last_messages[chat_id]
                    return send_new_message(chat_id, text, markup, parse_mode)
                raise
        return send_new_message(chat_id, text, markup, parse_mode)
    
    except Exception as e:
        logger.error(f"Ошибка при обновлении сообщения: {e}")
        return send_new_message(chat_id, text, markup, parse_mode)

def send_welcome_message(chat_id, user_data, weather_cache, message_id=None, force_new_message=False):
    markup = create_cities_keyboard(chat_id, user_data, weather_cache)
    
    user_cities = get_user_cities(chat_id, user_data)
    cities_weather_text = []
    
    for city in user_cities:
        cached_weather = get_cached_weather(city, weather_cache)
        if cached_weather:
            temp_str = f"+{cached_weather['temp']}" if cached_weather['temp'] > 0 else f"{cached_weather['temp']}"
            city_text = f"{cached_weather['emoji']} {city} {temp_str}°C"
            cities_weather_text.append(city_text)
        else:
            cities_weather_text.append(city)
    
    welcome_text = "\n".join(cities_weather_text) + "\n\nОтправь мне название населенного пункта и я скажу какая там погода и температура, дам советы по одежде.\n\n💡 Отправляй прогнозы в любой чат: введи @MeteoblueBot + город в любом чате Телеграм" if user_cities else "Отправь мне название населенного пункта и я скажу какая там погода и температура, дам советы по одежде.\n\n💡 Отправляй прогнозы в любой чат: введи @MeteoblueBot + город в любом чате Телеграм"
    
    if force_new_message:
        # Сначала удаляем предыдущее сообщение
        try:
            if message_id:
                bot.delete_message(chat_id, message_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить предыдущее сообщение: {e}")
        
        # Отправляем новое сообщение
        sent_message = send_new_message(chat_id, welcome_text, markup)
    else:
        # Обновляем существующее сообщение
        msg_info = last_messages.get(chat_id, {})
        msg_id = message_id if message_id else msg_info.get('message_id')
        
        sent_message = update_message(chat_id, msg_id, welcome_text, markup)
    
    if sent_message and hasattr(sent_message, 'message_id'):
        last_messages[chat_id] = {
            'message_id': sent_message.message_id,
            'timestamp': time.time()
        }

def send_reminder_message(chat_id):
    """Отправляет напоминание пользователям без городов"""
    reminder_text = (
        "🌤️ Привет! Пора узнать погоду на сегодня!\n\n"
        "Отправьте мне название города, чтобы получить актуальную информацию о погоде "
        "и рекомендации по одежде.\n\n"
        "Просто напишите название любого населенного пункта, и я расскажу:\n"
        "• Текущую температуру\n"
        "• Погодные условия\n"
        "• Что лучше надеть\n\n"
        "Попробуйте прямо сейчас! 😊"
    )
    
    try:
        sent_msg = bot.send_message(chat_id, reminder_text)
        last_messages[chat_id] = {
            'message_id': sent_msg.message_id,
            'timestamp': time.time()
        }
        return True
    except telebot.apihelper.ApiTelegramException as e:
        if "bot was blocked by the user" in str(e).lower():
            deactivate_user(chat_id)
            logger.info(f"Пользователь {chat_id} заблокировал бота. Деактивирован.")
            return False
    except Exception as e:
        logger.error(f"Ошибка при отправке напоминания пользователю {chat_id}: {e}")
        return False
    
# Функции для отслеживания активности
def initialize_excel_file():
    """Инициализирует Excel-файл для логирования, если он не существует"""
    try:
        if not os.path.exists(ACTIVITY_LOG_FILE):
            # Создаем DataFrame для вкладки с кнопкой "Обновить"
            df_refresh = pd.DataFrame(columns=['User ID', 'Count', 'Last Update'])
            
            # Создаем DataFrame для вкладки с городами
            df_cities = pd.DataFrame(columns=['User ID', 'City', 'Count', 'Last Update'])
            
            # Создаем DataFrame для вкладки с командой /start
            df_start = pd.DataFrame(columns=['User ID', 'Count', 'Last Update'])
            
            # Создаем DataFrame для вкладки с автообновлениями
            df_auto_updates = pd.DataFrame(columns=['Date', 'Time', 'Total Attempts', 'Messages Sent', 'Users With Cities', 'Users Without Cities', 'Blocked Users', 'Errors'])
            
            # Создаем writer для записи в Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl') as writer:
                df_refresh.to_excel(writer, sheet_name='Refresh Button', index=False)
                df_cities.to_excel(writer, sheet_name='City Buttons', index=False)
                df_start.to_excel(writer, sheet_name='Start Command', index=False)
                df_auto_updates.to_excel(writer, sheet_name='Auto Updates', index=False)
            
            activity_logger.info(f"Создан новый файл логов активности: {ACTIVITY_LOG_FILE}")
        else:
            # Проверяем файл на корректность
            try:
                workbook = load_workbook(ACTIVITY_LOG_FILE)
                
                # Проверяем наличие листа для команды /start
                if 'Start Command' not in workbook.sheetnames:
                    df_start = pd.DataFrame(columns=['User ID', 'Count', 'Last Update'])
                    with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a') as writer:
                        df_start.to_excel(writer, sheet_name='Start Command', index=False)
                    activity_logger.info(f"Добавлен лист 'Start Command' в существующий файл логов")
                
                # Проверяем наличие листа для автообновлений
                if 'Auto Updates' not in workbook.sheetnames:
                    df_auto_updates = pd.DataFrame(columns=['Date', 'Time', 'Total Attempts', 'Messages Sent', 'Users With Cities', 'Users Without Cities', 'Blocked Users', 'Errors'])
                    with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a') as writer:
                        df_auto_updates.to_excel(writer, sheet_name='Auto Updates', index=False)
                    activity_logger.info(f"Добавлен лист 'Auto Updates' в существующий файл логов")
                
                activity_logger.info(f"Существующий файл логов активности корректен: {ACTIVITY_LOG_FILE}")
            except InvalidFileException:
                activity_logger.error(f"Файл логов поврежден, создаем новый: {ACTIVITY_LOG_FILE}")
                os.remove(ACTIVITY_LOG_FILE)
                initialize_excel_file()
    except Exception as e:
        activity_logger.error(f"Ошибка при инициализации файла логов: {e}")

def update_refresh_log(user_id):
    """Обновляет информацию о нажатии кнопки 'Обновить'"""
    try:
        with excel_mutex:
            # Обновляем данные в памяти
            if user_id in activity_data['refresh']:
                activity_data['refresh'][user_id] += 1
            else:
                activity_data['refresh'][user_id] = 1
            
            # Текущая дата и время
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Загружаем данные из Excel
            df = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Refresh Button')
            
            # Проверяем, есть ли уже запись для данного пользователя
            user_row = df[df['User ID'] == user_id]
            
            if len(user_row) > 0:
                # Обновляем существующую запись
                df.loc[df['User ID'] == user_id, 'Count'] = activity_data['refresh'][user_id]
                df.loc[df['User ID'] == user_id, 'Last Update'] = current_time
            else:
                # Создаем новую запись
                new_row = pd.DataFrame({
                    'User ID': [user_id],
                    'Count': [activity_data['refresh'][user_id]],
                    'Last Update': [current_time]
                })
                df = pd.concat([df, new_row], ignore_index=True)
            
            # Сохраняем данные обратно в Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Refresh Button', index=False)
                
            activity_logger.info(f"Обновлен лог нажатия кнопки 'Обновить' для пользователя {user_id}")
    except Exception as e:
        activity_logger.error(f"Ошибка при обновлении лога нажатия кнопки 'Обновить': {e}")

def update_auto_updates_log(sent_count, with_cities_count, without_cities_count, blocked_count=0, error_count=0, total_attempts=0):
    """Обновляет информацию об автоматических обновлениях"""
    try:
        with excel_mutex:
            # Обновляем данные в памяти
            activity_data['auto_updates']['total_sent'] += sent_count
            activity_data['auto_updates']['last_update'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Текущая дата и время
            current_date = datetime.now().strftime('%Y-%m-%d')
            current_time = datetime.now().strftime('%H:%M:%S')
            
            # Загружаем данные из Excel
            df = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Auto Updates')
            
            # Добавляем новую запись
            new_row = pd.DataFrame({
                'Date': [current_date],
                'Time': [current_time],
                'Total Attempts': [total_attempts],
                'Messages Sent': [sent_count],
                'Users With Cities': [with_cities_count],
                'Users Without Cities': [without_cities_count],
                'Blocked Users': [blocked_count],
                'Errors': [error_count]
            })
            df = pd.concat([df, new_row], ignore_index=True)
            
            # Сохраняем данные обратно в Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Auto Updates', index=False)
                
            activity_logger.info(f"Обновлен лог автоматических обновлений: попыток {total_attempts}, отправлено {sent_count} сообщений (с городами: {with_cities_count}, без городов: {without_cities_count}, заблокировано: {blocked_count}, ошибок: {error_count})")
    except Exception as e:
        activity_logger.error(f"Ошибка при обновлении лога автоматических обновлений: {e}")

def auto_update_users():
    """Автоматически обновляет сообщения для всех пользователей каждые 4 часа"""
    while True:
        try:
            now = datetime.now()
            # Проверяем, наступило ли время обновления (каждые 4 часа: 00:01, 04:01, 08:01, 12:01, 16:01, 20:01)
            if now.minute == 1 and now.hour % 4 == 0:
                user_data = load_user_data()
                weather_cache = load_weather_cache()
                all_users = load_all_users()
                
                # Обновляем кэш погоды перед отправкой сообщений
                update_weather_cache(weather_cache, user_data)
                
                # Счетчики для подробной статистики
                total_attempts = 0
                sent_with_cities = 0
                sent_without_cities = 0
                blocked_count = 0
                error_count = 0
                total_sent = 0
                
                # Обрабатываем всех активных пользователей
                for user_id_str, user_info in all_users.items():
                    if not user_info.get('active', True):
                        continue
                        
                    total_attempts += 1
                    
                    try:
                        user_id = int(user_id_str)
                        user_cities = get_user_cities(user_id, user_data)
                        
                        if user_cities:
                            # Пользователь с городами - отправляем обновление погоды
                            msg_info = last_messages.get(user_id, {})
                            message_id = msg_info.get('message_id')
                            
                            send_welcome_message(user_id, user_data, weather_cache, message_id, force_new_message=True)
                            sent_with_cities += 1
                            total_sent += 1
                        else:
                            # Пользователь без городов - отправляем напоминание
                            if send_reminder_message(user_id):
                                sent_without_cities += 1
                                total_sent += 1
                        
                        # Небольшая задержка между отправками
                        time.sleep(0.1)
                        
                    except telebot.apihelper.ApiTelegramException as e:
                        if "bot was blocked by the user" in str(e).lower():
                            # Пользователь заблокировал бота
                            deactivate_user(user_id_str)
                            blocked_count += 1
                            logger.info(f"Пользователь {user_id_str} заблокировал бота. Деактивирован.")
                        else:
                            error_count += 1
                            logger.error(f"Ошибка при отправке автообновления пользователю {user_id_str}: {e}")
                    except Exception as e:
                        error_count += 1
                        logger.error(f"Ошибка при отправке автообновления пользователю {user_id_str}: {e}")
                
                # Логируем подробную статистику
                if total_attempts > 0:
                    update_auto_updates_log(total_sent, sent_with_cities, sent_without_cities, blocked_count, error_count, total_attempts)
                    
                    # Подробный лог
                    success_rate = (total_sent / total_attempts) * 100
                    logger.info(f"Автоматическое обновление в {now.strftime('%H:%M:%S')}:")
                    logger.info(f"  - Всего попыток: {total_attempts}")
                    logger.info(f"  - Успешно отправлено: {total_sent} ({success_rate:.1f}%)")
                    logger.info(f"  - С городами: {sent_with_cities}")
                    logger.info(f"  - Без городов: {sent_without_cities}")
                    logger.info(f"  - Заблокировали бота: {blocked_count}")
                    logger.info(f"  - Ошибки отправки: {error_count}")
                
                # Ждем 1 минуту, чтобы избежать повторного срабатывания
                time.sleep(60)
            
            # Проверяем каждые 30 секунд
            time.sleep(30)
            
        except Exception as e:
            logger.error(f"Ошибка в auto_update_users: {e}")
            time.sleep(300)  # 5 минут в случае ошибки

def update_start_log(user_id):
    """Обновляет информацию о использовании команды /start"""
    try:
        with excel_mutex:
            # Обновляем данные в памяти
            if user_id in activity_data['start']:
                activity_data['start'][user_id] += 1
            else:
                activity_data['start'][user_id] = 1
            
            # Текущая дата и время
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Загружаем данные из Excel
            df = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Start Command')
            
            # Проверяем, есть ли уже запись для данного пользователя
            user_row = df[df['User ID'] == user_id]
            
            if len(user_row) > 0:
                # Обновляем существующую запись
                df.loc[df['User ID'] == user_id, 'Count'] = activity_data['start'][user_id]
                df.loc[df['User ID'] == user_id, 'Last Update'] = current_time
            else:
                # Создаем новую запись
                new_row = pd.DataFrame({
                    'User ID': [user_id],
                    'Count': [activity_data['start'][user_id]],
                    'Last Update': [current_time]
                })
                df = pd.concat([df, new_row], ignore_index=True)
            
            # Сохраняем данные обратно в Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Start Command', index=False)
                
            activity_logger.info(f"Обновлен лог использования команды /start для пользователя {user_id}")
    except Exception as e:
        activity_logger.error(f"Ошибка при обновлении лога использования команды /start: {e}")

def update_city_log(user_id, city):
    """Обновляет информацию о нажатии кнопки с названием города"""
    try:
        with excel_mutex:
            # Обновляем данные в памяти
            if user_id not in activity_data['cities']:
                activity_data['cities'][user_id] = {}
                
            if city in activity_data['cities'][user_id]:
                activity_data['cities'][user_id][city] += 1
            else:
                activity_data['cities'][user_id][city] = 1
            
            # Текущая дата и время
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Загружаем данные из Excel
            df = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='City Buttons')
            
            # Проверяем, есть ли уже запись для данного пользователя и города
            user_city_row = df[(df['User ID'] == user_id) & (df['City'] == city)]
            
            if len(user_city_row) > 0:
                # Обновляем существующую запись
                df.loc[(df['User ID'] == user_id) & (df['City'] == city), 'Count'] = activity_data['cities'][user_id][city]
                df.loc[(df['User ID'] == user_id) & (df['City'] == city), 'Last Update'] = current_time
            else:
                # Создаем новую запись
                new_row = pd.DataFrame({
                    'User ID': [user_id],
                    'City': [city],
                    'Count': [activity_data['cities'][user_id][city]],
                    'Last Update': [current_time]
                })
                df = pd.concat([df, new_row], ignore_index=True)
            
            # Сохраняем данные обратно в Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='City Buttons', index=False)
                
            activity_logger.info(f"Обновлен лог нажатия кнопки города {city} для пользователя {user_id}")
    except Exception as e:
        activity_logger.error(f"Ошибка при обновлении лога нажатия кнопки города: {e}")

def generate_activity_report():
    """Генерирует отчет о активности пользователей"""
    try:
        with excel_mutex:
            # Загружаем данные из Excel
            df_refresh = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Refresh Button')
            df_cities = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='City Buttons')
            df_start = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Start Command')
            df_auto_updates = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Auto Updates')
            
            # Общая статистика
            total_refresh_clicks = df_refresh['Count'].sum() if not df_refresh.empty else 0
            total_city_clicks = df_cities['Count'].sum() if not df_cities.empty else 0
            total_start_cmds = df_start['Count'].sum() if not df_start.empty else 0
            total_auto_updates = df_auto_updates['Users Count'].sum() if not df_auto_updates.empty else 0
            
            # Загружаем данные о всех пользователях
            all_users = load_all_users()
            active_users = sum(1 for user in all_users.values() if user.get('active', True))
            
            # Топ-5 пользователей по нажатию кнопки "Обновить"
            top_refresh_users = df_refresh.sort_values('Count', ascending=False).head(5) if not df_refresh.empty else pd.DataFrame()
            
            # Топ-5 пользователей по использованию команды /start
            top_start_users = df_start.sort_values('Count', ascending=False).head(5) if not df_start.empty else pd.DataFrame()
            
            # Топ-5 самых популярных городов
            if not df_cities.empty and 'City' in df_cities.columns:
                top_cities = df_cities.groupby('City')['Count'].sum().reset_index().sort_values('Count', ascending=False).head(5)
            else:
                top_cities = pd.DataFrame()
            
            # Последние 5 автоматических обновлений
            if not df_auto_updates.empty:
                last_updates = df_auto_updates.tail(5)
                last_updates_str = "\n".join([
                    f"{row['Date']} {row['Time']} - Попыток: {row.get('Total Attempts', 'N/A')}, "
                    f"Отправлено: {row.get('Messages Sent', row.get('Users Count', 'N/A'))}, "
                    f"С городами: {row.get('Users With Cities', 'N/A')}, "
                    f"Без городов: {row.get('Users Without Cities', 'N/A')}, "
                    f"Заблокировано: {row.get('Blocked Users', 'N/A')}, "
                    f"Ошибок: {row.get('Errors', 'N/A')}"
                    for _, row in last_updates.iterrows()
                ])
            else:
                last_updates_str = "Нет данных"
            
            # Формируем отчет
            report = f"Отчет о активности пользователей (на {datetime.now().strftime('%Y-%m-%d %H:%M:%S')})\n\n"
            report += f"Всего зарегистрированных пользователей: {len(all_users)}\n"
            report += f"Активных пользователей: {active_users}\n"
            report += f"Общее количество нажатий на кнопку 'Обновить': {total_refresh_clicks}\n"
            report += f"Общее количество нажатий на кнопки городов: {total_city_clicks}\n"
            report += f"Общее количество использований команды /start: {total_start_cmds}\n"
            report += f"Общее количество автоматических обновлений: {total_auto_updates}\n\n"
            
            report += "Последние 5 автоматических обновлений:\n" + last_updates_str + "\n\n"
            
            if not top_refresh_users.empty:
                report += "Топ-5 пользователей по нажатиям кнопки 'Обновить':\n"
                for idx, row in top_refresh_users.iterrows():
                    report += f"- Пользователь {row['User ID']}: {row['Count']} нажатий\n"
                report += "\n"
            
            if not top_start_users.empty:
                report += "Топ-5 пользователей по использованию команды /start:\n"
                for idx, row in top_start_users.iterrows():
                    report += f"- Пользователь {row['User ID']}: {row['Count']} использований\n"
                report += "\n"
            
            if not top_cities.empty:
                report += "Топ-5 самых популярных городов:\n"
                for idx, row in top_cities.iterrows():
                    report += f"- {row['City']}: {row['Count']} нажатий\n"
            
            # Создаем отчет в Excel
            with pd.ExcelWriter(ACTIVITY_LOG_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Добавляем вкладку с отчетом
                workbook = writer.book
                if 'Activity Report' in workbook.sheetnames:
                    workbook.remove(workbook['Activity Report'])
                
                worksheet = workbook.create_sheet('Activity Report')
                worksheet['A1'] = report
                worksheet['A1'].font = Font(size=12)
                worksheet['A1'].alignment = Alignment(wrap_text=True)
                
                # Устанавливаем ширину колонки
                worksheet.column_dimensions['A'].width = 100
                
            activity_logger.info(f"Сгенерирован отчет о активности пользователей")
            return report
    except Exception as e:
        activity_logger.error(f"Ошибка при генерации отчета о активности: {e}")
        return f"Ошибка при генерации отчета: {e}"

def load_activity_data():
    """Загружает данные активности из Excel в память"""
    try:
        if os.path.exists(ACTIVITY_LOG_FILE):
            # Загружаем данные кнопки "Обновить"
            df_refresh = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Refresh Button')
            for idx, row in df_refresh.iterrows():
                activity_data['refresh'][row['User ID']] = row['Count']
            
            # Загружаем данные кнопок городов
            df_cities = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='City Buttons')
            for idx, row in df_cities.iterrows():
                user_id = row['User ID']
                city = row['City']
                count = row['Count']
                
                if user_id not in activity_data['cities']:
                    activity_data['cities'][user_id] = {}
                
                activity_data['cities'][user_id][city] = count
            
            # Загружаем данные команды /start
            try:
                df_start = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Start Command')
                for idx, row in df_start.iterrows():
                    activity_data['start'][row['User ID']] = row['Count']
            except Exception as e:
                activity_logger.warning(f"Не удалось загрузить данные о команде /start: {e}")
            
            # Загружаем данные автообновлений
            try:
                df_auto_updates = pd.read_excel(ACTIVITY_LOG_FILE, sheet_name='Auto Updates')
                if not df_auto_updates.empty:
                    if 'Messages Sent' in df_auto_updates.columns:
                        activity_data['auto_updates']['total_sent'] = df_auto_updates['Messages Sent'].sum()
                    elif 'Users Count' in df_auto_updates.columns:
                        activity_data['auto_updates']['total_sent'] = df_auto_updates['Users Count'].sum()
                    
                    last_update = df_auto_updates.iloc[-1]
                    activity_data['auto_updates']['last_update'] = f"{last_update['Date']} {last_update['Time']}"
            except Exception as e:
                activity_logger.warning(f"Не удалось загрузить данные об автообновлениях: {e}")
                
            activity_logger.info(f"Загружены данные активности из {ACTIVITY_LOG_FILE}")
    except Exception as e:
        activity_logger.error(f"Ошибка при загрузке данных активности: {e}")

# Обработчик команды /start
@bot.message_handler(commands=['start'])
def start(message):
    try:
        chat_id = message.chat.id
        
        # Добавляем пользователя в общий список
        add_user_to_all_users(chat_id)
        
        # Отслеживаем использование команды /start
        update_start_log(chat_id)
        
        if chat_id in last_messages:
            del last_messages[chat_id]
            
        user_data = load_user_data()
        weather_cache = load_weather_cache()
        send_welcome_message(chat_id, user_data, weather_cache)
        
    except Exception as e:
        logger.error(f"Ошибка в обработчике /start: {e}")
        bot.send_message(message.chat.id, "Давайте начнем заново! Вот меню:")
        user_data = load_user_data()
        weather_cache = load_weather_cache()
        send_welcome_message(message.chat.id, user_data, weather_cache)

# Обработчик команды /stats для генерации отчета активности
@bot.message_handler(commands=['stats'])
def stats(message):
    try:
        # Проверяем, что запрос от администратора (можно настроить список админов)
        admin_ids = [message.chat.id]  # Здесь можно добавить список ID админов
        
        if message.chat.id in admin_ids:
            report = generate_activity_report()
            bot.send_message(message.chat.id, f"Отчет об активности пользователей:\n\n{report}")
        else:
            bot.send_message(message.chat.id, "У вас нет доступа к этой команде.")
    except Exception as e:
        logger.error(f"Ошибка в обработчике /stats: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка при генерации отчета.")

# Обработчик команды /check_users для проверки статуса пользователей
@bot.message_handler(commands=['check_users'])
def check_users(message):
    try:
        # Проверяем, что запрос от администратора
        admin_ids = [message.chat.id]  # Здесь можно добавить список ID админов
        
        if message.chat.id in admin_ids:
            all_users = load_all_users()
            user_data = load_user_data()
            
            active_users = 0
            inactive_users = 0
            users_with_cities = 0
            users_without_cities = 0
            additional_list_users = 0
            start_command_users = 0
            
            for user_id_str, user_info in all_users.items():
                if user_info.get('active', True):
                    active_users += 1
                    user_cities = get_user_cities(int(user_id_str), user_data)
                    
                    if user_cities:
                        users_with_cities += 1
                    else:
                        users_without_cities += 1
                else:
                    inactive_users += 1
                
                # Проверяем источник пользователя
                source = user_info.get('source', 'unknown')
                if source == 'additional_list':
                    additional_list_users += 1
                elif source == 'start_command':
                    start_command_users += 1
            
            check_report = f"Проверка статуса пользователей (на {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}):\n\n"
            check_report += f"📊 Общая статистика:\n"
            check_report += f"• Всего пользователей: {len(all_users)}\n"
            check_report += f"• Активных: {active_users}\n"
            check_report += f"• Неактивных (заблокировали бота): {inactive_users}\n\n"
            
            check_report += f"🏙️ По наличию городов:\n"
            check_report += f"• С добавленными городами: {users_with_cities}\n"
            check_report += f"• Без городов: {users_without_cities}\n\n"
            
            check_report += f"📝 По источникам:\n"
            check_report += f"• Из дополнительного списка: {additional_list_users}\n"
            check_report += f"• Через команду /start: {start_command_users}\n\n"
            
            success_rate = (active_users / len(all_users)) * 100 if len(all_users) > 0 else 0
            check_report += f"✅ Процент активных пользователей: {success_rate:.1f}%\n"
            
            bot.send_message(message.chat.id, check_report)
        else:
            bot.send_message(message.chat.id, "У вас нет доступа к этой команде.")
    except Exception as e:
        logger.error(f"Ошибка в обработчике /check_users: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка при проверке пользователей.")

def get_and_send_weather(chat_id, city, user_data, weather_cache, message_id=None, force_new_message=False):
    url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&units=metric&lang=ru&appid={config.OPENWEATHER_API_KEY}'
    
    try:
        weather_data = requests.get(url).json()
        
        if weather_data.get('cod') == '404':
            if force_new_message and message_id:
                try:
                    bot.delete_message(chat_id, message_id)
                except Exception as e:
                    logger.warning(f"Не удалось удалить предыдущее сообщение: {e}")
                
                error_msg = send_new_message(chat_id, 'Город не найден. Проверьте написание.')
            else:
                msg_info = last_messages.get(chat_id, {})
                msg_id = message_id if message_id else msg_info.get('message_id')
                error_msg = update_message(chat_id, msg_id, 'Город не найден. Проверьте написание.')
            
            if error_msg and hasattr(error_msg, 'message_id'):
                last_messages[chat_id] = {
                    'message_id': error_msg.message_id,
                    'timestamp': time.time()
                }
            return False
            
        temperature = round(weather_data['main']['temp'], 2)
        temperature_feels = round(weather_data['main']['feels_like'], 2)
        weather_description = weather_data['weather'][0]['description']
        wind_speed = weather_data['wind']['speed']
        
        # Определение текущего времени года и времени суток
        from datetime import datetime
        current_date = datetime.now()
        
        # Определяем время года
        month = current_date.month
        if month in [12, 1, 2]:
            season = 'зима'
        elif month in [3, 4, 5]:
            season = 'весна'
        elif month in [6, 7, 8]:
            season = 'лето'
        else:  # 9, 10, 11
            season = 'осень'
        
        # Определяем время суток
        hour = current_date.hour
        if 6 <= hour < 12:
            time_of_day = 'утро'
        elif 12 <= hour < 18:
            time_of_day = 'день'
        elif 18 <= hour < 24:
            time_of_day = 'вечер'
        else:  # 0 <= hour < 6
            time_of_day = 'ночь'
        
        weather_emoji = get_weather_emoji(weather_description)
        # Используем обновленную функцию с учетом времени года и времени суток
        clothes_advice = get_clothing_advice(temperature, weather_description, season, time_of_day, wind_speed)
        
        temp_str = f"+{temperature}" if temperature > 0 else f"{temperature}"
        temp_feels_str = f"+{temperature_feels}" if temperature_feels > 0 else f"{temperature_feels}"
        weather_description_cap = weather_description[0].upper() + weather_description[1:]
        
        current_time = int(time.time())
        formatted_time = format_date_time(current_time)
        
        weather_message = (
            f'{weather_emoji} {city} {weather_description_cap}\n'
            f'🌡️ t° {temp_str}°C\n'
            f'🌡️ t°ощущ. {temp_feels_str}°C\n'
            f'💨 Скорость ветра | {wind_speed} м/с\n'
            f'{clothes_advice}\n'
            f'⏱️ Время обновления: {formatted_time}'
        )
        
        cache_data = {
            'temp': temperature,
            'emoji': weather_emoji,
            'description': weather_description,
            'wind_speed': wind_speed,
            'updated_at': current_time
        }
        update_cached_weather(city, cache_data, weather_cache)
        
        markup = types.InlineKeyboardMarkup(row_width=2)
        user_cities = get_user_cities(chat_id, user_data)
        
        if city in user_cities:
            markup.add(
                types.InlineKeyboardButton("Удалить город", callback_data=f"remove_{city}"),
                types.InlineKeyboardButton("Назад", callback_data="back")
            )
        else:
            markup.add(
                types.InlineKeyboardButton("Добавить город", callback_data=f"add_{city}"),
                types.InlineKeyboardButton("Назад", callback_data="back")
            )
        
        if force_new_message:
            # Сначала удаляем предыдущее сообщение
            try:
                if message_id:
                    bot.delete_message(chat_id, message_id)
            except Exception as e:
                logger.warning(f"Не удалось удалить предыдущее сообщение: {e}")
            
            # Отправляем новое сообщение
            sent_message = send_new_message(chat_id, weather_message, markup)
        else:
            # Обновляем существующее сообщение
            msg_info = last_messages.get(chat_id, {})
            msg_id = message_id if message_id else msg_info.get('message_id')
            sent_message = update_message(chat_id, msg_id, weather_message, markup)
        
        if sent_message and hasattr(sent_message, 'message_id'):
            last_messages[chat_id] = {
                'message_id': sent_message.message_id,
                'timestamp': time.time()
            }
            
        return True
    
    except Exception as e:
        logger.error(f'Ошибка при получении погоды: {e}')
        
        if force_new_message and message_id:
            try:
                bot.delete_message(chat_id, message_id)
            except Exception as e:
                logger.warning(f"Не удалось удалить предыдущее сообщение: {e}")
            
            error_msg = send_new_message(
                chat_id, 
                'Произошла ошибка при получении данных о погоде. Попробуйте позже.'
            )
        else:
            msg_info = last_messages.get(chat_id, {})
            msg_id = message_id if message_id else msg_info.get('message_id')
            
            error_msg = update_message(
                chat_id, 
                msg_id, 
                'Произошла ошибка при получении данных о погоде. Попробуйте позже.'
            )
        
        if error_msg and hasattr(error_msg, 'message_id'):
            last_messages[chat_id] = {
                'message_id': error_msg.message_id,
                'timestamp': time.time()
            }
            
        return False

@bot.callback_query_handler(func=lambda call: True)
def callback_handler(call):
    user_data = load_user_data()
    weather_cache = load_weather_cache()
    
    try:
        if call.data == "back":
            send_welcome_message(call.message.chat.id, user_data, weather_cache, call.message.message_id)
        
        elif call.data == "refresh":
            # Отслеживаем нажатие кнопки "Обновить"
            update_refresh_log(call.message.chat.id)
            
            # Обновляем данные о погоде для всех городов пользователя
            user_cities = get_user_cities(call.message.chat.id, user_data)
            for city in user_cities:
                get_weather_data(city, weather_cache)
                time.sleep(0.5)  # Небольшая задержка между запросами
            
            # Отправляем новое сообщение, но сначала удаляем предыдущее
            send_welcome_message(call.message.chat.id, user_data, weather_cache, call.message.message_id, force_new_message=True)
        
        elif call.data.startswith("add_"):
            city = call.data[4:]
            if add_user_city(call.message.chat.id, city, user_data):
                send_welcome_message(call.message.chat.id, user_data, weather_cache, call.message.message_id)
        
        elif call.data.startswith("remove_"):
            city = call.data[7:]
            if remove_user_city(call.message.chat.id, city, user_data):
                send_welcome_message(call.message.chat.id, user_data, weather_cache, call.message.message_id)
        
        elif call.data.startswith("city_"):
            city = call.data[5:]
            # Отслеживаем нажатие кнопки города
            update_city_log(call.message.chat.id, city)
            
            get_and_send_weather(call.message.chat.id, city, user_data, weather_cache, call.message.message_id, force_new_message=False)
            
        bot.answer_callback_query(call.id)
    except Exception as e:
        logger.error(f"Ошибка в callback_handler: {e}")
        try:
            bot.answer_callback_query(call.id)
        except:
            pass

@bot.message_handler(func=lambda message: True)
def handle_all_messages(message):
    try:
        if message.text.startswith('/'):
            bot.send_message(message.chat.id, "Неизвестная команда. Попробуйте /start")
        else:
            user_data = load_user_data()
            weather_cache = load_weather_cache()
            get_and_send_weather(message.chat.id, message.text, user_data, weather_cache, last_messages.get(message.chat.id, {}).get('message_id'), force_new_message=False)
        
        try:
            bot.delete_message(message.chat.id, message.message_id)
        except Exception as e:
            logger.warning(f"Не удалось удалить сообщение пользователя: {e}")
    except Exception as e:
        logger.error(f"Ошибка обработки сообщения: {e}")
        bot.send_message(message.chat.id, "Произошла ошибка. Попробуйте позже.")

@bot.inline_handler(func=lambda query: True)
def handle_inline_query(query):
    try:
        city = query.query.strip()
        
        if not city:
            # Если город не указан, предлагаем варианты с актуальной погодой
            # Получаем данные о погоде для Москвы
            moscow_weather = get_weather_data("Москва", load_weather_cache())
            if not moscow_weather:
                moscow_temp = "+14"
                moscow_temp_feels = "+12"
                moscow_description = "Облачно с прояснениями"
                moscow_wind = 6
                moscow_emoji = "⛅"
            else:
                moscow_temp = f"+{moscow_weather['temp']}" if moscow_weather['temp'] > 0 else f"{moscow_weather['temp']}"
                moscow_temp_feels = f"+{moscow_weather['temp'] - 2}" if moscow_weather['temp'] > 0 else f"{moscow_weather['temp'] - 2}"
                moscow_description = "Облачно с прояснениями"
                moscow_wind = moscow_weather['wind_speed']
                moscow_emoji = moscow_weather['emoji']
            
            # Получаем данные о погоде для Санкт-Петербурга
            spb_weather = get_weather_data("Санкт-Петербург", load_weather_cache())
            if not spb_weather:
                spb_temp = "+8"
                spb_temp_feels = "+3"
                spb_description = "Облачно"
                spb_wind = 5
                spb_emoji = "☁️"
            else:
                spb_temp = f"+{spb_weather['temp']}" if spb_weather['temp'] > 0 else f"{spb_weather['temp']}"
                spb_temp_feels = f"+{spb_weather['temp'] - 5}" if spb_weather['temp'] > 0 else f"{spb_weather['temp'] - 5}"
                spb_description = "Облачно"
                spb_wind = spb_weather['wind_speed']
                spb_emoji = spb_weather['emoji']
            
            # Определение текущего времени года и времени суток
            from datetime import datetime
            current_date = datetime.now()
            
            # Определяем время года
            month = current_date.month
            if month in [12, 1, 2]:
                season = 'зима'
            elif month in [3, 4, 5]:
                season = 'весна'
            elif month in [6, 7, 8]:
                season = 'лето'
            else:  # 9, 10, 11
                season = 'осень'
            
            # Определяем время суток
            hour = current_date.hour
            if 6 <= hour < 12:
                time_of_day = 'утро'
            elif 12 <= hour < 18:
                time_of_day = 'день'
            elif 18 <= hour < 24:
                time_of_day = 'вечер'
            else:  # 0 <= hour < 6
                time_of_day = 'ночь'
            
            # Формируем советы по одежде с учетом времени года и времени суток
            moscow_clothes_advice = get_clothing_advice(float(moscow_temp.replace("+", "")), moscow_description, season, time_of_day, moscow_wind)
            spb_clothes_advice = get_clothing_advice(float(spb_temp.replace("+", "")), spb_description, season, time_of_day, spb_wind)
            
            current_time = int(time.time())
            formatted_time = format_date_time(current_time)
            
            results = [
                types.InlineQueryResultArticle(
                    id="1",
                    title="Москва",
                    description=f"{moscow_temp}°C, {moscow_description}",
                    input_message_content=types.InputTextMessageContent(
                        message_text=f"{moscow_emoji} Москва {moscow_description}\n🌡️ t° {moscow_temp}°C\n🌡️ t°ощущ. {moscow_temp_feels}°C\n💨 Скорость ветра | {moscow_wind} м/с\n {moscow_clothes_advice}\n⏱️ Время обновления: {formatted_time}"
                    )
                ),
                types.InlineQueryResultArticle(
                    id="2",
                    title="Санкт-Петербург",
                    description=f"{spb_temp}°C, {spb_description}",
                    input_message_content=types.InputTextMessageContent(
                        message_text=f"{spb_emoji} Санкт-Петербург {spb_description}\n🌡️ t° {spb_temp}°C\n🌡️ t°ощущ. {spb_temp_feels}°C\n💨 Скорость ветра | {spb_wind} м/с\n {spb_clothes_advice}\n⏱️ Время обновления: {formatted_time}"
                    )
                )
            ]
            bot.answer_inline_query(query.id, results)
            return

        # Получаем реальные данные о погоде
        url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&units=metric&lang=ru&appid={config.OPENWEATHER_API_KEY}'
        response = requests.get(url)
        response.raise_for_status()
        weather_data = response.json()
        
        if weather_data.get('cod') == '404':
            bot.answer_inline_query(query.id, [])
            return
            
        temperature = round(weather_data['main']['temp'], 2)
        temperature_feels = round(weather_data['main']['feels_like'], 2)
        weather_description = weather_data['weather'][0]['description']
        wind_speed = weather_data['wind']['speed']
        
        # Определение текущего времени года и времени суток
        from datetime import datetime
        current_date = datetime.now()
        
        # Определяем время года
        month = current_date.month
        if month in [12, 1, 2]:
            season = 'зима'
        elif month in [3, 4, 5]:
            season = 'весна'
        elif month in [6, 7, 8]:
            season = 'лето'
        else:  # 9, 10, 11
            season = 'осень'
        
        # Определяем время суток
        hour = current_date.hour
        if 6 <= hour < 12:
            time_of_day = 'утро'
        elif 12 <= hour < 18:
            time_of_day = 'день'
        elif 18 <= hour < 24:
            time_of_day = 'вечер'
        else:  # 0 <= hour < 6
            time_of_day = 'ночь'
        
        weather_emoji = get_weather_emoji(weather_description)
        clothes_advice = get_clothing_advice(temperature, weather_description, season, time_of_day, wind_speed)
        
        temp_str = f"+{temperature}" if temperature > 0 else f"{temperature}"
        temp_feels_str = f"+{temperature_feels}" if temperature_feels > 0 else f"{temperature_feels}"
        weather_description_cap = weather_description[0].upper() + weather_description[1:]
        
        current_time = int(time.time())
        formatted_time = format_date_time(current_time)
        
        # Формируем сообщение в точном соответствии с примером
        message_text = (
            f"{weather_emoji} {city} {weather_description_cap}\n"
            f"🌡️ t° {temp_str}°C\n"
            f"🌡️ t°ощущ. {temp_feels_str}°C\n"
            f"💨 Скорость ветра | {wind_speed} м/с\n"
            f"{clothes_advice}\n"
            f"⏱️ Время обновления: {formatted_time}"
        )

        result = types.InlineQueryResultArticle(
            id="1",
            title=f"Погода в {city}",
            description=f"{temp_str}°C, {weather_description}",
            input_message_content=types.InputTextMessageContent(
                message_text=message_text
            )
        )

        bot.answer_inline_query(query.id, [result])

    except Exception as e:
        logger.error(f"Inline error: {e}")
        bot.answer_inline_query(query.id, [])

def clean_message_cache():
    while True:
        try:
            current_time = time.time()
            to_delete = [chat_id for chat_id, msg_info in last_messages.items()
                        if current_time - msg_info.get('timestamp', 0) > 604800]

            for chat_id in to_delete:
                del last_messages[chat_id]

            time.sleep(86400)
        except Exception as e:
            logger.error(f"Ошибка в clean_message_cache: {e}")
            time.sleep(3600)

# Flask сервер для Railway (keepalive)
app = Flask(__name__)

@app.route('/')
def home():
    return "GidMeteo Bot is running!", 200

@app.route('/health')
def health():
    return "OK", 200

def run_flask():
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)

def update_all_weather_info():
    while True:
        try:
            logger.info(f"Обновление кэша погоды... {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            user_data = load_user_data()
            weather_cache = load_weather_cache()
            update_weather_cache(weather_cache, user_data)
            
            logger.info(f"Обновление кэша погоды завершено! {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            time.sleep(600)
        except Exception as e:
            logger.error(f"Ошибка при обновлении кэша: {e}")
            time.sleep(60)

if __name__ == '__main__':
    if not os.path.exists(USER_DATA_FILE):
        save_user_data({})
    
    if not os.path.exists(WEATHER_CACHE_FILE):
        save_weather_cache({})
    
    if not os.path.exists(ALL_USERS_FILE):
        save_all_users({})
    
    # Добавляем пользователей из списка в базу данных
    added_users = add_additional_users_to_all_users()
    
    # Инициализируем Excel-файл для отслеживания активности
    initialize_excel_file()
    
    # Загружаем данные активности
    load_activity_data()
    
    # Запускаем потоки для фоновых задач
    threading.Thread(target=clean_message_cache, daemon=True).start()
    threading.Thread(target=update_all_weather_info, daemon=True).start()
    threading.Thread(target=auto_update_users, daemon=True).start()

    # Запускаем Flask сервер в отдельном потоке (для Railway)
    if os.environ.get('PORT'):
        threading.Thread(target=run_flask, daemon=True).start()
        logger.info(f"Flask сервер запущен на порту {os.environ.get('PORT')}")

    logger.info("Бот GidMeteo запущен")
    print("Бот GidMeteo запущен. Нажмите Ctrl+C для остановки.")
    print("Отслеживание активности включено. Данные сохраняются в файл:", ACTIVITY_LOG_FILE)
    print("Автоматические обновления будут отправляться каждые 4 часа (00:01, 04:01, 08:01, 12:01, 16:01, 20:01)")
    print(f"Добавлено {added_users} новых пользователей из дополнительного списка")
    print("Доступные команды для администратора:")
    print("  /stats - подробный отчет активности")
    print("  /check_users - проверка статуса всех пользователей")
    
    try:
        while True:
            try:
                bot.polling(none_stop=True, interval=0, timeout=60)
            except Exception as e:
                logger.error(f'Сработало исключение: {e}')
                time.sleep(10)
    except KeyboardInterrupt:
        logger.info("Получен сигнал Ctrl+C. Останавливаю бота...")
        print("\nБот остановлен.")